from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import FileResponse, HTMLResponse, StreamingResponse
import asyncio, threading, queue
import pandas as pd
import shutil, os, uuid, re, base64, pathlib
from ortools.sat.python import cp_model
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

app = FastAPI(title="Smart Shift by OR-Tools")
TEMP_DIR = "temp_files"
os.makedirs(TEMP_DIR, exist_ok=True)

WORK_SHIFTS = ["早", "遅", "夜", "日"]
REST_SHIFTS  = ["×", "有", "○", "公", "△"]   # ○=夜勤明け日休み, ×=土休み, 公=公休(32h用), △=standalone日休み(土扱い)
ALL_SHIFTS   = WORK_SHIFTS + REST_SHIFTS

PINK_FILL   = PatternFill("solid", fgColor="FFB6C1")
GREEN_FILL  = PatternFill("solid", fgColor="90EE90")
YELLOW_FILL = PatternFill("solid", fgColor="FFFF99")
GRAY_FILL   = PatternFill("solid", fgColor="D3D3D3")
BLUE_FILL   = PatternFill("solid", fgColor="BDD7EE")   # 主任使用日
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FILL2= PatternFill("solid", fgColor="5B9BD5")
A_UNIT_FILL = PatternFill("solid", fgColor="DEEAF1")   # Aユニット薄青
B_UNIT_FILL = PatternFill("solid", fgColor="E2EFDA")   # Bユニット薄緑

WEEKDAY_MAP = {
    "月": 0, "火": 1, "水": 2, "木": 3, "金": 4, "土": 5, "日": 6,
    "月曜": 0, "火曜": 1, "水曜": 2, "木曜": 3, "金曜": 4, "土曜": 5, "日曜": 6,
}


# ========================================================
# Settings 読み込み
# ========================================================
def load_settings(df):
    start, end = None, None
    holidays = {}
    nenkyuu = {}
    nikkin_days = []
    holiday_periods = []
    
    try:
        for r in [1]: 
            for c in [3, 4]: 
                val = str(df.iloc[r, c]).strip()
                day_map = {"月曜":0, "火曜":1, "水曜":2, "木曜":3, "金曜":4, "土曜":5, "日曜":6}
                if val in day_map:
                    nikkin_days.append(day_map[val])
    except:
        pass

    header_row = None
    for i in range(len(df)):
        v = str(df.iloc[i, 0]).strip()
        if "期間" in v and "開始" in v:
            header_row = i
            break
    if header_row is None:
        raise Exception("Settingsシートに期間ヘッダーが見つかりません")

    for j in range(header_row + 1, len(df)):
        s_val = pd.to_datetime(df.iloc[j, 0], errors="coerce")
        e_val = pd.to_datetime(df.iloc[j, 1], errors="coerce")
        c = str(df.iloc[j, 2]).strip()
        # D列=土, E列=日, F列=公, G列=年休（新列構造）
        n_土_str  = str(df.iloc[j, 3]).strip() if df.shape[1] > 3 else ""
        n_日_str  = str(df.iloc[j, 4]).strip() if df.shape[1] > 4 else ""
        n_公_str  = str(df.iloc[j, 5]).strip() if df.shape[1] > 5 else ""
        try:
            nen_str = str(df.iloc[j, 6]).strip() if df.shape[1] > 6 else ""
        except Exception:
            nen_str = ""
        if pd.isna(s_val) and pd.isna(e_val) and c in ["nan", "None", ""]:
            continue
        if pd.notna(s_val):
            start = s_val if start is None else min(start, s_val)
        if pd.notna(e_val):
            end = e_val if end is None else max(end, e_val)

        m_土 = re.search(r"\d+", n_土_str)
        m_日 = re.search(r"\d+", n_日_str)
        m_公 = re.search(r"\d+", n_公_str)
        if (m_土 or m_日 or m_公) and c not in ["nan", "None", ""]:
            n_土 = int(m_土.group()) if m_土 else 0
            n_日 = int(m_日.group()) if m_日 else 0
            n_公 = int(m_公.group()) if m_公 else 0
            num  = n_土 + n_日 + n_公
            p_start_dt = s_val.to_pydatetime().replace(tzinfo=None, hour=0, minute=0, second=0, microsecond=0) if pd.notna(s_val) else None
            p_end_dt   = e_val.to_pydatetime().replace(tzinfo=None, hour=0, minute=0, second=0, microsecond=0) if pd.notna(e_val) else None
            if "40" in c:
                holidays["40h"] = holidays.get("40h", 0) + num
                if p_start_dt and p_end_dt:
                    holiday_periods.append((p_start_dt, p_end_dt, "40h", n_土, n_日, n_公))
            elif "32" in c:
                holidays["32h"] = holidays.get("32h", 0) + num
                if p_start_dt and p_end_dt:
                    holiday_periods.append((p_start_dt, p_end_dt, "32h", n_土, n_日, n_公))
            elif "パート" in c:
                holidays["パート"] = holidays.get("パート", 0) + num
                if p_start_dt and p_end_dt:
                    holiday_periods.append((p_start_dt, p_end_dt, "パート", n_土, n_日, n_公))
                
        nen_m = re.search(r"\d+", nen_str)
        if nen_m and c not in ["nan", "None", ""]:
            nen_num = int(nen_m.group())
            if "40" in c:
                nenkyuu["40h"] = nenkyuu.get("40h", 0) + nen_num
            elif "32" in c:
                nenkyuu["32h"] = nenkyuu.get("32h", 0) + nen_num
            elif "パート" in c:
                nenkyuu["パート"] = nenkyuu.get("パート", 0) + nen_num

    if start is None or end is None:
        raise Exception("期間が取得できませんでした")

    if not holidays:
        holidays = {"40h": 9, "32h": 8, "パート": 0}
    else:
        holidays.setdefault("40h", 0)
        holidays.setdefault("32h", 0)
        holidays.setdefault("パート", 0)

    if not nenkyuu:
        nenkyuu = {"40h": 2, "32h": 2, "パート": 0}
    else:
        nenkyuu.setdefault("40h", 0)
        nenkyuu.setdefault("32h", 0)
        nenkyuu.setdefault("パート", 0)

    days = []
    d = start
    while d <= end:
        days.append(d)
        d += timedelta(days=1)
    return days, holidays, nenkyuu, nikkin_days, holiday_periods


# ========================================================
# 希望シフト 読み込み
# ========================================================
def load_requests(df, days, staff_list, part_staff=None):
    if part_staff is None:
        part_staff = []
    requests = {}

    header_row = None
    for i in range(len(df)):
        if str(df.iloc[i, 0]).strip() == "職員名":
            header_row = i
            break
    if header_row is None:
        return requests

    col_to_date = {}
    for j in range(1, len(df.columns)):
        d = pd.to_datetime(df.iloc[header_row, j], errors="coerce")
        if pd.notna(d):
            col_to_date[j] = d.to_pydatetime().replace(
                tzinfo=None, hour=0, minute=0, second=0, microsecond=0)

    data_start = header_row + 2
    for i in range(data_start, len(df)):
        name = str(df.iloc[i, 0]).strip()
        if name in ["nan", "None", "", "0"] or name not in staff_list:
            continue
        requests[name] = {}
        is_part = (name in part_staff)
        for j, date in col_to_date.items():
            raw = str(df.iloc[i, j]).strip()
            if raw in ["nan", "None", "", "0"]:
                continue
            if "×" in raw or "土" in raw or "休み" in raw:
                requests[name][date] = ("×", "希望")
            elif raw == "公" or "公休" in raw:
                requests[name][date] = ("公", "希望")
            elif "有給" in raw or raw in ("有", "年") or "年休" in raw:
                requests[name][date] = ("有", "指定" if is_part else "希望")
            elif "夜勤" in raw or raw == "夜":
                requests[name][date] = ("夜", "指定")
            elif "早出" in raw or raw in ("早", "ハ"):
                requests[name][date] = ("早", "指定")
            elif "遅出" in raw or raw in ("遅", "オ"):
                requests[name][date] = ("遅", "指定")
            elif "日勤" in raw or raw == "ニ":
                requests[name][date] = ("日", "指定")
            elif raw == "日":
                # 新システムでは「日」リクエスト=standalone日休み(△)。日勤はニで入力。
                requests[name][date] = ("△", "希望")
    return requests


# ========================================================
# 前月実績 読み込み
# ========================================================
def load_prev_month(df, staff_list):
    prev = {}
    header_row = None
    for i in range(len(df)):
        if str(df.iloc[i, 0]).strip() == "職員名":
            header_row = i
            break
    if header_row is None:
        return prev

    date_cols = []
    for j in range(1, len(df.columns)):
        d = pd.to_datetime(df.iloc[header_row, j], errors="coerce")
        if pd.notna(d):
            date_cols.append(j)

    for i in range(header_row + 1, len(df)):
        name = str(df.iloc[i, 0]).strip()
        if name in ["nan", "None", "", "0"] or name not in staff_list:
            continue
        seq = []
        for j in date_cols:
            raw = str(df.iloc[i, j]).strip()
            if "夜勤" in raw or raw == "夜":                               seq.append("夜")
            elif "早出" in raw or raw in ("早", "ハ"):                     seq.append("早")
            elif "遅出" in raw or raw in ("遅", "オ"):                     seq.append("遅")
            elif "日勤" in raw or raw == "ニ":                             seq.append("日")
            elif "有給" in raw or raw in ("有", "年") or "年休" in raw:    seq.append("有")
            elif raw == "日":
                # 直前が夜勤なら夜勤明け(○)、そうでなければstandalone日休み(△)
                seq.append("○" if (seq and seq[-1] == "夜") else "△")
            elif raw == "公":                                               seq.append("公")
            elif raw == "土" or raw == "×":                                seq.append("×")
            else:                                                           seq.append("×")
        prev[name] = seq
    return prev


def count_trailing_consec(shift_seq):
    count = 0
    for s in reversed(shift_seq):
        if s in ["早", "遅", "夜", "日", "有"]:
            count += 1
        else:
            break
    return count


# ========================================================
# INFEASIBLE 診断ヘルパー
# ========================================================
def _diagnose_infeasible(staff, shuunin_list, requests, days_norm, N,
                         allowed_shifts_map, fixed_holiday_map, holiday_limits,
                         cont_map, nmin_map, nmax_map, prev_month, weekly_work_days,
                         unit_map=None, ab_staff_set=None,
                         weekday_allowed_map=None, nikkin_days_settings=None,
                         ojt_list=None):
    msgs = []
    seen = set()
    _ojt = set(ojt_list or [])
    # OJTも診断対象に含める（staffはそのまま使用）

    def add_msg(m):
        if m not in seen:
            seen.add(m)
            msgs.append(m)

    SHIFT_NAME = {
        "早": "早出(ハ)", "遅": "遅出(オ)", "日": "日勤(ニ)",
        "夜": "夜勤", "×": "土休み(土)", "有": "年休(有)", "○": "日休み(日/夜勤明け)", "公": "公休(公)", "△": "日休み(日/単独)"
    }

    for s in staff:
        allowed = allowed_shifts_map.get(s)
        if allowed is None:
            continue
        forbidden_reqs = []
        for date_obj, (sh_type, req_type) in requests.get(s, {}).items():
            if sh_type not in ["早","遅","日","夜"]:
                continue
            is_in_allowed = sh_type in allowed
            if req_type == "指定":
                continue 
            if not is_in_allowed and req_type == "希望":
                forbidden_reqs.append(
                    f"{date_obj.strftime('%m/%d')}({SHIFT_NAME.get(sh_type,sh_type)})希望")
        if forbidden_reqs:
            allowed_names = ",".join(SHIFT_NAME.get(a, a) for a in sorted(allowed))
            days_str = ", ".join(forbidden_reqs)
            add_msg(
                f"警告: {s}さんの 希望シフトが備考制限と矛盾しています（希望は無視されます）。\n"
                f"  備考制限: 許可勤務は [{allowed_names}] のみ\n"
                f"  矛盾する希望: {days_str}\n"
                f"  → 備考制限を変更するか、Shift_Requestsシートで指定（指定勤務）に変更してください。"
            )

    for s in shuunin_list:
        bad_reqs = []
        for date_obj, (sh_type, req_type) in requests.get(s, {}).items():
            if sh_type in ["遅","夜"] and req_type == "指定":
                continue 
            if sh_type in ["日","有","○","公","△"] and req_type == "指定":
                bad_reqs.append(
                    f"{date_obj.strftime('%m/%d')}({SHIFT_NAME.get(sh_type,sh_type)})指定")
        if bad_reqs:
            days_str = ", ".join(bad_reqs)
            add_msg(
                f"致命的エラー: {s}さん（主任）への指定が矛盾しています。\n"
                f"  主任は 早出(ハ)・遅出(オ)・夜勤・土休み(土) のみ指定可能です。\n"
                f"  矛盾する指定: {days_str}\n"
                f"  → Shift_Requestsシートで該当日を早出(ハ)・遅出(オ)・夜勤・土休み(土)に変更してください。"
            )

    for s in staff:
        hol_limit = holiday_limits.get(cont_map.get(s, "40h"), 0)
        if hol_limit == 0:
            continue
        hope_off_days = sum(
            1 for _date_obj, (sh_type, req_type) in requests.get(s, {}).items()
            if sh_type in ["×", "有", "公", "△"] and req_type == "希望"
        )
        if hope_off_days > hol_limit:
            add_msg(
                f"致命的エラー: {s}さん の 希望休の数({hope_off_days}日) が "
                f"設定公休数({hol_limit}日) を超えているため、スケジュールを確定できません。\n"
                f"  → 希望休を{hol_limit}日以内に絞るか、公休数設定を見直してください。"
            )

    for s in staff:
        hol = holiday_limits.get(cont_map.get(s, "40h"), 0)
        nmin = nmin_map.get(s, 0)
        nmax = nmax_map.get(s, 0)
        prev_seq = prev_month.get(s, [])
        prev_night = (bool(prev_seq) and prev_seq[-1] == "夜")
        max_maru = nmax + (1 if prev_night else 0)
        if hol > 0 and max_maru > hol:
            add_msg(
                f"致命的エラー: {s}さん の 夜勤上限回数と公休数のバランスが取れていません"
                f"（夜勤の翌日は日休み(日)になるため、勤務できる日数が足りません）。\n"
                f"  公休{hol}日 に対し 夜勤上限{nmax}回"
                + ("＋前月末夜勤(+1)" if prev_night else "")
                + f" = 日(夜勤明け){max_maru}日必要ですが、{max_maru - hol}日不足します。\n"
                f"  → 夜勤上限を{hol - (1 if prev_night else 0)}以下にするか、"
                f"公休数を{max_maru}以上に増やしてください。"
            )
        if nmin > N:
            add_msg(
                f"致命的エラー: {s}さんの 夜勤最少数({nmin}回) が 対象日数({N}日) を超えています。\n"
                f"  → 夜勤最少数を{N}以下に設定してください。"
            )

    night_capable = [s for s in staff if nmax_map.get(s, 0) > 0]
    total_nmax = sum(nmax_map.get(s, 0) for s in night_capable)
    total_nmin = sum(nmin_map.get(s, 0) for s in night_capable)
    if total_nmin > N:
        names = ", ".join(f"{s}({nmin_map[s]}回)" for s in night_capable)
        add_msg(
            f"致命的エラー: 夜勤可能スタッフ全員の夜勤最少数合計({total_nmin}回) が "
            f"対象日数({N}日) を超えています。\n"
            f"  [{names}]\n"
            f"  → 夜勤最少数の合計が{N}以下になるよう各職員の設定を見直してください。"
        )
    if total_nmax < N:
        names = ", ".join(f"{s}({nmax_map[s]}回)" for s in night_capable)
        add_msg(
            f"致命的エラー: 夜勤可能スタッフ全員の夜勤最高数合計({total_nmax}回) が "
            f"対象日数({N}日) より少ないです。\n"
            f"  [{names}]\n"
            f"  → 夜勤最高数の合計が{N}以上になるよう各職員の設定を見直してください。"
        )

    if unit_map is not None and days_norm:
        _ab_set = ab_staff_set or set()
        for d, dn in enumerate(days_norm):
            date_str = f"{dn.month}月{dn.day}日"
            a_avail = []
            b_avail = []
            for s in staff:
                req = requests.get(s, {}).get(dn)
                fixed_off = (dn.weekday() in (fixed_holiday_map.get(s) or set()))
                is_req_off = (req and req[0] in ["×", "有"] and req[1] == "指定")
                if fixed_off or is_req_off:
                    continue
                u = unit_map.get(s, "")
                is_ab = (s in _ab_set)
                if u == "A" or is_ab:
                    a_avail.append(s)
                if u == "B" or is_ab:
                    b_avail.append(s)
            if len(a_avail) < 1:
                add_msg(
                    f"致命的エラー: {date_str} の Aユニット・早出 の必要人数(1名)に対して、"
                    f"出勤可能なスタッフが不足しています。\n"
                    f"  → {date_str}前後の希望休・固定休を見直してください。"
                )
            if len(b_avail) < 1:
                add_msg(
                    f"致命的エラー: {date_str} の Bユニット・早出 の必要人数(1名)に対して、"
                    f"出勤可能なスタッフが不足しています。\n"
                    f"  → {date_str}前後の希望休・固定休を見直してください。"
                )

    if weekday_allowed_map:
        for s, wd_map in weekday_allowed_map.items():
            for wd, allowed_wd in wd_map.items():
                for d, dn in enumerate(days_norm):
                    if dn.weekday() == wd:
                        req = requests.get(s, {}).get(dn)
                        if req and req[1] == "希望" and req[0] not in allowed_wd:
                            allowed_names = ",".join(SHIFT_NAME.get(a, a) for a in sorted(allowed_wd))
                            add_msg(
                                f"警告: {s}さんの {dn.strftime('%m/%d')} の希望({SHIFT_NAME.get(req[0],req[0])})が"
                                f"備考の曜日指定({allowed_names})と矛盾しています。\n"
                                f"  → 備考の曜日指定を変更するか、希望を修正してください。"
                            )

    if nikkin_days_settings and days_norm:
        for wd_target in nikkin_days_settings:
            for d, dn in enumerate(days_norm):
                if dn.weekday() == wd_target:
                    nikkin_avail = []
                    for s in staff:
                        req = requests.get(s, {}).get(dn)
                        if req and req[1] == "指定" and req[0] != "日":
                            continue
                        allowed = allowed_shifts_map.get(s)
                        if allowed is not None and "日" not in allowed:
                            continue
                        if s in weekday_allowed_map and wd_target in weekday_allowed_map[s]:
                            if "日" not in weekday_allowed_map[s][wd_target]:
                                continue
                        nikkin_avail.append(s)
                    
                    if not nikkin_avail:
                        day_name = ["月","火","水","木","金","土","日"][wd_target]
                        add_msg(
                            f"致命的エラー: {dn.strftime('%m/%d')}({day_name}) の日勤配置(1名)に対して、"
                            f"出勤可能なスタッフが不足しています。\n"
                            f"  → 該当日付近の希望休や備考の勤務制限を見直してください。"
                        )

    for s in [s2 for s2 in staff if cont_map.get(s2) == "パート"]:
        warn_days = []
        for date_obj, (sh_type, req_type) in requests.get(s, {}).items():
            if sh_type == "有" and req_type == "希望":
                warn_days.append(date_obj.strftime("%m/%d"))
        if warn_days:
            add_msg(
                f"警告: {s}さん（パート）の年休(有)希望は無視されます。\n"
                f"  対象日: {', '.join(warn_days)}\n"
                f"  → '指定'に変更するか、空白にしてください。"
            )

    # ─── 日付単位の詳細衝突診断 ───────────────────────────────────────
    if unit_map is not None and days_norm:
        _ab_set = ab_staff_set or set()
        OFF_SHIFTS = {"×", "有", "公", "△", "○"}
        WD_JA = ["月","火","水","木","金","土","日"]

        def is_off_on(s, dn):
            req = requests.get(s, {}).get(dn)
            if req and req[0] in OFF_SHIFTS:
                return True
            if dn.weekday() in (fixed_holiday_map.get(s) or set()):
                return True
            return False

        def off_reason(s, dn):
            req = requests.get(s, {}).get(dn)
            if req and req[0] in OFF_SHIFTS:
                return "希望休" if req[1] == "希望" else "指定休"
            if dn.weekday() in (fixed_holiday_map.get(s) or set()):
                return "固定休"
            return "休み"

        def can_work_shift(s, dn, sh):
            if is_off_on(s, dn):
                return False
            req = requests.get(s, {}).get(dn)
            if req and req[1] == "指定" and req[0] != sh:
                return False
            allowed = (allowed_shifts_map or {}).get(s)
            if allowed is not None and sh not in allowed:
                return False
            wd = dn.weekday()
            if s in (weekday_allowed_map or {}) and wd in weekday_allowed_map[s]:
                if sh not in weekday_allowed_map[s][wd]:
                    return False
            return True

        for d, dn in enumerate(days_norm):
            date_label = f"{dn.month}月{dn.day}日({WD_JA[dn.weekday()]})"

            # A早出
            a_early = [s for s in staff if (unit_map.get(s) == "A" or s in _ab_set) and can_work_shift(s, dn, "早")]
            shuunin_ok = [s for s in shuunin_list if not is_off_on(s, dn)]
            if not a_early and not shuunin_ok:
                off_list = [f"{s}（{off_reason(s,dn)}）" for s in staff if (unit_map.get(s) == "A" or s in _ab_set) and is_off_on(s, dn)]
                detail = f"\n  休み中: {', '.join(off_list)}" if off_list else ""
                add_msg(f"致命的エラー: {date_label} は Aユニット早出に配置できるスタッフがいません。{detail}\n  → 希望休・固定休を見直してください。")

            # B早出
            b_early = [s for s in staff if (unit_map.get(s) == "B" or s in _ab_set) and can_work_shift(s, dn, "早")]
            if not b_early and not shuunin_ok:
                off_list = [f"{s}（{off_reason(s,dn)}）" for s in staff if (unit_map.get(s) == "B" or s in _ab_set) and is_off_on(s, dn)]
                detail = f"\n  休み中: {', '.join(off_list)}" if off_list else ""
                add_msg(f"致命的エラー: {date_label} は Bユニット早出に配置できるスタッフがいません。{detail}\n  → 希望休・固定休を見直してください。")

            # A遅出
            a_late = [s for s in staff if (unit_map.get(s) == "A" or s in _ab_set) and can_work_shift(s, dn, "遅")]
            if not a_late:
                off_list = [f"{s}（{off_reason(s,dn)}）" for s in staff if (unit_map.get(s) == "A" or s in _ab_set) and is_off_on(s, dn)]
                detail = f"\n  休み中: {', '.join(off_list)}" if off_list else ""
                add_msg(f"致命的エラー: {date_label} は A遅出に配置できるスタッフがいません。{detail}\n  → 希望休・固定休を見直してください。")

            # B遅出
            b_late = [s for s in staff if (unit_map.get(s) == "B" or s in _ab_set) and can_work_shift(s, dn, "遅")]
            if not b_late:
                off_list = [f"{s}（{off_reason(s,dn)}）" for s in staff if (unit_map.get(s) == "B" or s in _ab_set) and is_off_on(s, dn)]
                detail = f"\n  休み中: {', '.join(off_list)}" if off_list else ""
                add_msg(f"致命的エラー: {date_label} は B遅出に配置できるスタッフがいません。{detail}\n  → 希望休・固定休を見直してください。")

            # 夜勤
            night_avail = [s for s in staff if nmax_map.get(s, 0) > 0 and can_work_shift(s, dn, "夜")]
            if not night_avail:
                off_list = [f"{s}（{off_reason(s,dn)}）" for s in staff if nmax_map.get(s, 0) > 0 and is_off_on(s, dn)]
                detail = f"\n  夜勤可能スタッフの休み: {', '.join(off_list)}" if off_list else ""
                add_msg(f"致命的エラー: {date_label} は夜勤に配置できるスタッフがいません。{detail}\n  → 希望休・固定休を見直してください。")

    # ─── OJT固有の診断 ──────────────────────────────────────────────
    if _ojt and unit_map is not None and days_norm:
        WD_JA = ["月","火","水","木","金","土","日"]
        OFF_SHIFTS_D = {"×", "有", "公", "△", "○"}

        for s in _ojt:
            instr = (ojt_list or []) and None  # dummy init
        # ojt_instructorは渡されていないのでrequestsから類推するか、
        # staffのholiday_limitsで判定する

        # OJTの希望休チェック（通常スタッフと同様に）
        for s in _ojt:
            hol_limit = holiday_limits.get(cont_map.get(s, "40h"), 0)
            if hol_limit == 0:
                continue
            hope_off = sum(1 for _, (sh, rt) in requests.get(s, {}).items()
                           if sh in ["×","有","公","△"] and rt == "希望")
            if hope_off > hol_limit:
                add_msg(
                    f"致命的エラー(OJT): {s}さん(OJT) の希望休({hope_off}日)が"
                    f"公休数({hol_limit}日)を超えています。\n"
                    f"  → Shift_Requestsで希望休を{hol_limit}日以内に絞ってください。"
                )

        # OJTに割り当てられた指導者が存在するか
        # （ojt_instructorは診断関数に渡されていないため、
        #   requests・note等から間接的に検出できる場合のみ警告）
        for s in _ojt:
            # 固定公休と希望の衝突チェック
            for d, dn in enumerate(days_norm):
                date_label = f"{dn.month}月{dn.day}日({WD_JA[dn.weekday()]})"
                req = requests.get(s, {}).get(dn)
                fixed_off = (dn.weekday() in (fixed_holiday_map.get(s) or set()))
                if fixed_off and req and req[1] == "指定" and req[0] not in OFF_SHIFTS_D:
                    add_msg(
                        f"致命的エラー(OJT): {s}さん(OJT) の {date_label} は固定公休ですが、"
                        f"指定勤務({req[0]})が入っています。\n"
                        f"  → 固定公休か指定勤務のどちらかを取り消してください。"
                    )

    return msgs

# ========================================================
# 計算進捗管理
# ========================================================
_progress_queues: dict = {}  # uid -> queue.Queue
_result_cache: dict   = {}  # uid -> preview JSON dict（生成結果キャッシュ）

def _cache_preview(uid: str, result, staff, shuunin_list, unit_map, cont_map,
                   role_map, days_norm, requests, ab_unit_result, shuunin_unit_result,
                   kanmu_map, prev_month, nmin_map, nmax_map, consec_night_map, holiday_periods):
    """生成結果からプレビューJSONを作ってキャッシュする"""
    try:
        DISPLAY_MAP = {"×": "土", "○": "日", "公": "公", "△": "日"}
        SHIFT_ABBR  = {"早": "ハ", "遅": "オ", "日": "ニ", "有": "年"}

        def disp(s, d):
            sh = result.get(s, {}).get(d, "×")
            if sh in DISPLAY_MAP: return DISPLAY_MAP[sh]
            if sh == "日": return "ニ"
            if sh == "有": return "年"
            if sh not in ("早", "遅"): return sh
            abbr = SHIFT_ABBR[sh]
            if s in shuunin_list:
                unit = shuunin_unit_result.get(s, {}).get(d)
                return (unit + abbr) if unit else abbr
            elif kanmu_map.get(s, "×") == "○":
                unit = ab_unit_result.get(s, {}).get(d)
                return (unit + abbr) if unit else abbr
            else:
                unit = unit_map.get(s, "")
                return (unit + abbr) if unit in ("A", "B") else abbr

        def req_type(s, d):
            dn = days_norm[d]
            req = requests.get(s, {}).get(dn)
            return req[1] if req else ""

        WD = ["月","火","水","木","金","土","日"]
        days_info = [{"day": dn.day, "wd": WD[dn.weekday()],
                      "is_sat": dn.weekday()==5, "is_sun": dn.weekday()==6}
                     for dn in days_norm]

        all_staff = shuunin_list + staff
        rows = []
        for s in all_staff:
            cells = [{"v": disp(s, d), "r": req_type(s, d)} for d in range(len(days_norm))]
            rows.append({"name": s, "unit": unit_map.get(s, "主任"), "cells": cells})

        total_score, total_max, all_rows = score_shift(
            result, staff, shuunin_list, days_norm, requests,
            prev_month, cont_map, role_map, nmin_map, nmax_map,
            consec_night_map, holiday_periods, unit_map,
            ab_unit_result, shuunin_unit_result, kanmu_map)

        deduct_items = [{"name": n, "count": c, "per": p, "total": t}
                        for n, c, p, t in all_rows if t > 0]

        _result_cache[uid] = {
            "days": days_info, "rows": rows,
            "score": total_score, "score_max": total_max,
            "deductions": deduct_items
        }
    except Exception:
        pass  # キャッシュ失敗はサイレントに無視


def _get_progress_queue(uid: str) -> queue.Queue:
    q = queue.Queue()
    _progress_queues[uid] = q
    return q

def _push_progress(uid: str, msg: str):
    if uid and uid in _progress_queues:
        _progress_queues[uid].put(msg)

def _close_progress(uid: str):
    if uid and uid in _progress_queues:
        _progress_queues[uid].put(None)  # sentinel
        _progress_queues.pop(uid, None)


class ProgressCallback(cp_model.CpSolverSolutionCallback):
    """OR-Tools が解を見つけるたびに呼ばれるコールバック"""
    def __init__(self, uid: str, N: int, phase: str = ""):
        super().__init__()
        self._uid = uid
        self._N = N
        self._phase = phase
        self._count = 0
        self._start = None

    def on_solution_callback(self):
        import time
        if self._start is None:
            self._start = time.time()
        self._count += 1
        elapsed = time.time() - self._start
        obj = int(self.ObjectiveValue())
        msg = (f"{self._phase}解 #{self._count} 発見 "
               f"| スコア目標値: {obj} "
               f"| 経過: {elapsed:.1f}秒")
        _push_progress(self._uid, msg)


def generate_shift(file_path, random_seed=0, validate_only=False, timeout=300, progress_uid=None, num_workers=8):
    xls = pd.ExcelFile(file_path)
    staff_df    = xls.parse("Staff_Master",   header=None)
    settings_df = xls.parse("Settings",       header=None)
    request_df  = xls.parse("Shift_Requests", header=None)
    prev_df     = xls.parse("Prev_Month",     header=None)

    # ── 通常スタッフ（最初の「職員名」ヘッダーまで） ──────────────────
    first_header_row = None
    for i in range(len(staff_df)):
        if str(staff_df.iloc[i, 0]).strip() == "職員名":
            first_header_row = i
            staff_df.columns = staff_df.iloc[i]
            staff_df = staff_df.iloc[i+1:].reset_index(drop=True)
            break

    # ── OJT スタッフ（2つ目の「職員名」ヘッダーを探す） ───────────────
    # staff_df には既に1行目ヘッダーが除去されているので、残りから探す
    ojt_df = None
    ojt_header_idx = None
    for i in range(len(staff_df)):
        val = str(staff_df.iloc[i, 0]).strip()
        if val == "職員名":
            ojt_header_idx = i
            break

    if ojt_header_idx is not None:
        # OJT セクションを切り出し
        ojt_block = staff_df.iloc[ojt_header_idx:].copy()
        ojt_block.columns = ojt_block.iloc[0]
        ojt_block = ojt_block.iloc[1:].reset_index(drop=True)
        ojt_block = ojt_block[ojt_block["職員名"].notna()].copy()
        ojt_block = ojt_block[~ojt_block["職員名"].astype(str).isin(["nan","0",""])].copy()
        ojt_block["職員名"] = ojt_block["職員名"].astype(str).str.strip()
        # 通常スタッフ側からOJTセクションを除去
        staff_df = staff_df.iloc[:ojt_header_idx].copy()
        ojt_df = ojt_block

    staff_df = staff_df[staff_df["職員名"].notna()].copy()
    staff_df = staff_df[~staff_df["職員名"].astype(str).isin(["nan","0",""])].copy()
    staff_df["職員名"] = staff_df["職員名"].astype(str).str.strip()

    def col_num(name, default=0):
        if name in staff_df.columns:
            col_data = staff_df[name]
            if isinstance(col_data, pd.DataFrame):
                col_data = col_data.iloc[:, 0]
            return pd.to_numeric(col_data, errors="coerce").fillna(default).astype(int)
        return pd.Series([default]*len(staff_df))

    staff_df["夜勤最少数"] = col_num("夜勤最少数", 0)
    staff_df["夜勤最高数"] = col_num("夜勤最高数", 0)

    all_staff_names = staff_df["職員名"].tolist()

    def get_map(col, default=""):
        if col in staff_df.columns:
            return dict(zip(staff_df["職員名"], staff_df[col].astype(str).str.strip()))
        return {s: default for s in all_staff_names}

    unit_map  = get_map("ユニット")
    cont_map  = get_map("契約区分")
    role_map  = get_map("役職")
    nmin_map  = dict(zip(staff_df["職員名"], staff_df["夜勤最少数"]))
    nmax_map  = dict(zip(staff_df["職員名"], staff_df["夜勤最高数"]))
    note_map  = get_map("備考")

    # 前半夜勤NG（K列）: ○ の人は月の前半（15日以前）に夜勤を入れない
    zenhan_ng_map = get_map("前半夜勤NG", default="×")
    for k in zenhan_ng_map:
        if zenhan_ng_map[k] in ["", "nan", "None"]:
            zenhan_ng_map[k] = "×"

    # 連続夜勤（J列追加対応）
    consec_night_map = get_map("連続夜勤希望", default="")
    # 古いフォーマット用フォールバック
    if all(v in ["", "nan", "None", "×"] for v in consec_night_map.values()):
        consec_night_map = get_map("連続夜勤", default="×")
    for k in consec_night_map:
        if consec_night_map[k] in ["", "nan", "None"]:
            consec_night_map[k] = "×"

    kanmu_col = next((c for c in staff_df.columns if "兼務" in str(c)), None)
    if kanmu_col:
        kanmu_map = dict(zip(staff_df["職員名"], staff_df[kanmu_col].astype(str).str.strip()))
    else:
        kanmu_map = {}
        for s in all_staff_names:
            u = str(unit_map.get(s, "")).strip()
            if u == "A・B":
                kanmu_map[s] = "○"
            else:
                kanmu_map[s] = "×"

    fixed_holiday_map = {}
    fhcol = next((c for c in staff_df.columns if "固定" in str(c) and "休" in str(c)), None)
    if fhcol:
        for _, row in staff_df.iterrows():
            val = str(row[fhcol]).strip()
            if val in ["nan","None","","-","0"]:
                continue
            wdays = [WEEKDAY_MAP[t.strip()] for t in re.split(r"[,、・\s]+", val)
                     if t.strip() in WEEKDAY_MAP]
            if wdays:
                fixed_holiday_map[row["職員名"]] = wdays

    shuunin_list = [s for s in all_staff_names
                    if str(unit_map.get(s, "")).lower() in ("nan", "", "none")]
    staff = [s for s in all_staff_names if s not in shuunin_list]
    part_staff = [s for s in staff if cont_map.get(s) == "パート"]

    # ── OJT スタッフの処理 ─────────────────────────────────────────────
    ojt_list = []       # OJT スタッフ名リスト
    ojt_instructor = {} # ojt名 → 指導者名

    # OJT の契約区分を holiday_limits の既知キーに正規化するマッピング
    # 「特定技能」「研修生」等 → "32h" か "40h" に近い方へ正規化
    VALID_CONT_KEYS = {"40h", "32h", "パート"}
    def normalize_cont(ct):
        if ct in VALID_CONT_KEYS:
            return ct
        # 週32h以下っぽいキーワードなら32h、それ以外は40h
        if any(kw in ct for kw in ["32", "パート", "短", "特定", "研修", "OJT"]):
            return "32h"
        return "40h"

    if ojt_df is not None and len(ojt_df) > 0:
        def ojt_col_num(df, name, default=0):
            if name in df.columns:
                col_data = df[name]
                if isinstance(col_data, pd.DataFrame):
                    col_data = col_data.iloc[:, 0]
                return pd.to_numeric(col_data, errors="coerce").fillna(default).astype(int)
            return pd.Series([default]*len(df))

        ojt_df["夜勤最少数"] = ojt_col_num(ojt_df, "夜勤最少数", 0)
        ojt_df["夜勤最高数"] = ojt_col_num(ojt_df, "夜勤最高数", 0)

        for _, row in ojt_df.iterrows():
            s = str(row["職員名"]).strip()
            if not s or s in ("nan", "None", "0"):
                continue
            ojt_list.append(s)
            # 指導者列（E列: "指導者" または "ユニット兼務"）
            instr_col = next((c for c in ojt_df.columns if "指導" in str(c)), None)
            if instr_col is None:
                instr_col = next((c for c in ojt_df.columns if "兼務" in str(c)), None)
            instr = str(row[instr_col]).strip() if instr_col else ""
            if instr not in ("nan", "None", ""):
                ojt_instructor[s] = instr
            # マップに追加（契約区分は正規化して登録）
            unit_map[s]  = str(row.get("ユニット", "")).strip()
            raw_cont     = str(row.get("契約区分", "40h")).strip()
            cont_map[s]  = normalize_cont(raw_cont)
            role_map[s]  = str(row.get("役職", "")).strip()
            nmin_map[s]  = int(row["夜勤最少数"])
            nmax_map[s]  = int(row["夜勤最高数"])
            note_map[s]  = str(row.get("備考", "")).strip()
            if note_map[s] in ("nan", "None"): note_map[s] = ""
            kanmu_map[s] = "×"
            zenhan_ng_map[s] = str(row.get("前半夜勤NG", "×")).strip()
            if zenhan_ng_map[s] in ("nan", "None", ""): zenhan_ng_map[s] = "×"
            cn_val = str(row.get("連続夜勤希望", "×")).strip()
            if cn_val in ("nan", "None", ""): cn_val = "×"
            consec_night_map[s] = cn_val
            # 固定公休
            fhcol2 = next((c for c in ojt_df.columns if "固定" in str(c) and "休" in str(c)), None)
            if fhcol2:
                fh_val = str(row.get(fhcol2, "")).strip()
                if fh_val not in ("nan", "None", "", "-", "0"):
                    wdays = [WEEKDAY_MAP[t.strip()] for t in re.split(r"[,、・\s]+", fh_val)
                             if t.strip() in WEEKDAY_MAP]
                    if wdays:
                        fixed_holiday_map[s] = wdays
            all_staff_names.append(s)

        # OJT は staff リストに追加（ただし配置カウント対象外・part_staffには含めない）
        staff.extend(ojt_list)

    days, holiday_limits, nenkyuu_limits, nikkin_days_settings, holiday_periods = load_settings(settings_df)
    N = len(days)
    all_names_for_req = all_staff_names
    requests   = load_requests(request_df, days, all_names_for_req, part_staff=part_staff)
    prev_month = load_prev_month(prev_df, all_names_for_req)

    def to_naive(d):
        if hasattr(d, 'to_pydatetime'):
            return d.to_pydatetime().replace(tzinfo=None, hour=0, minute=0, second=0, microsecond=0)
        return datetime(d.year, d.month, d.day)
    days_norm = [to_naive(d) for d in days]

    allowed_shifts_map = {}
    weekly_work_days   = {}
    weekday_allowed_map = {}
    part_with_fixed = set()
    
    # 備考の曜日指定解析強化版
    shift_map_full = {"早出":"早", "早":"早", "遅出":"遅", "遅":"遅", "夜勤":"夜", "夜":"夜", "日勤":"日", "日":"日", "休み":"×", "×":"×", "有給":"有", "年休":"有", "有":"有"}

    for s in all_staff_names:
        note = str(note_map.get(s, ""))
        allowed = None
        if "早出のみ" in note:
            allowed = {"早"}
        elif "遅出のみ" in note:
            allowed = {"遅"}
        elif "夜勤なし" in note or "夜勤禁止" in note:
            allowed = {"早", "遅", "日"}
        if allowed is not None:
            allowed_shifts_map[s] = allowed

        pattern = r"(月|火|水|木|金|土|日)曜?[は：:]+([^。]+)"
        for m in re.finditer(pattern, note):
            wd_str = m.group(1)
            rule_str = m.group(2)
            wd = {"月":0, "火":1, "水":2, "木":3, "金":4, "土":5, "日":6}[wd_str]

            allowed_wd = set()
            all_possible = {"早", "遅", "夜", "日", "×", "有"}
            
            if "以外" in rule_str:
                forbidden_wd = set()
                for k, v in shift_map_full.items():
                    if k in rule_str.split("以外")[0]:
                        forbidden_wd.add(v)
                allowed_wd = all_possible - forbidden_wd
            else:
                for k, v in shift_map_full.items():
                    if k in rule_str:
                        allowed_wd.add(v)
                if not allowed_wd:
                    continue
                # 指定の勤務に限定された場合でも、その日の休み(×・有)は許可する
                allowed_wd.add("×")
                allowed_wd.add("有")
                
            if allowed_wd:
                if s not in weekday_allowed_map:
                    weekday_allowed_map[s] = {}
                weekday_allowed_map[s][wd] = allowed_wd

        m = re.search(r"週(\d+)日", note)
        if m:
            weekly_work_days[s] = int(m.group(1))

    for s in part_staff:
        req_s = requests.get(s, {})
        designated = sum(1 for v in req_s.values() if v[1] == "指定" and v[0] in WORK_SHIFTS)
        if designated >= 3:
            part_with_fixed.add(s)

    week_groups = defaultdict(list)
    for d_idx, dn in enumerate(days_norm):
        sun_offset = (dn.weekday() + 1) % 7
        week_sun   = dn - timedelta(days=sun_offset)
        week_groups[week_sun.strftime("%Y-%m-%d")].append(d_idx)
    sorted_week_keys = sorted(week_groups.keys())

    ab_staff = [s for s in staff if kanmu_map.get(s, "×") == "○"]
    ab_staff_set = set(ab_staff)

    model = cp_model.CpModel()

    x = {}
    for s in staff:
        for d in range(N):
            for sh in ALL_SHIFTS:
                x[s, d, sh] = model.NewBoolVar(f"x_{s}_{d}_{sh}")

    xs = {}
    for s in shuunin_list:
        for d in range(N):
            for sh in ALL_SHIFTS:
                xs[s, d, sh] = model.NewBoolVar(f"xs_{s}_{d}_{sh}")

    uea = {}; ueb = {}; ula = {}; ulb = {}
    for s in ab_staff:
        for d in range(N):
            uea[s,d] = model.NewBoolVar(f"uea_{s}_{d}")
            ueb[s,d] = model.NewBoolVar(f"ueb_{s}_{d}")
            ula[s,d] = model.NewBoolVar(f"ula_{s}_{d}")
            ulb[s,d] = model.NewBoolVar(f"ulb_{s}_{d}")
            model.Add(uea[s,d] + ueb[s,d] == x[s,d,"早"])
            model.Add(ula[s,d] + ulb[s,d] == x[s,d,"遅"])

    shuunin_use_a = {}; shuunin_use_b = {}
    for s in shuunin_list:
        for d in range(N):
            shuunin_use_a[s,d] = model.NewBoolVar(f"sh_ua_{s}_{d}")
            shuunin_use_b[s,d] = model.NewBoolVar(f"sh_ub_{s}_{d}")
            model.Add(shuunin_use_a[s,d] + shuunin_use_b[s,d] <= xs[s,d,"早"])
            model.Add(shuunin_use_a[s,d] + shuunin_use_b[s,d] <= 1)

    for s in staff:
        for d in range(N):
            model.AddExactlyOne(x[s,d,sh] for sh in ALL_SHIFTS)
    for s in shuunin_list:
        for d in range(N):
            model.AddExactlyOne(xs[s,d,sh] for sh in ALL_SHIFTS)

    def fix_requests(var_dict, s_list):
        for s in s_list:
            if s not in requests:
                continue
            for date_obj, (sh_type, req_type) in requests[s].items():
                for d, dn in enumerate(days_norm):
                    if dn == date_obj and sh_type in ALL_SHIFTS:
                        model.Add(var_dict[s,d,sh_type] == 1)
                        break
    fix_requests(x, staff)
    fix_requests(xs, shuunin_list)

    for s in staff:
        if prev_month.get(s, []) and prev_month[s][-1] == "夜":
            for sh_f in ["早","遅","日","夜","×"]:
                model.Add(x[s,0,sh_f] == 0)
    for s in shuunin_list:
        if prev_month.get(s, []) and prev_month[s][-1] == "夜":
            model.Add(xs[s,0,"○"] == 1)

    for s, wdays in fixed_holiday_map.items():
        var_dict = xs if s in shuunin_list else x
        for d_idx, dn in enumerate(days_norm):
            if dn.weekday() in wdays:
                req = requests.get(s, {}).get(dn)
                if req and req[1] == "指定":
                    continue
                model.Add(var_dict[s,d_idx,"×"] == 1)

    # OJT は配置カウントに含めない（非OJTのみで制約を構成）
    non_ojt_staff = [s for s in staff if s not in ojt_list]

    for d in range(N):
        a_e = ([x[s,d,"早"] for s in non_ojt_staff if unit_map.get(s) == "A" and s not in ab_staff_set] +
               [uea[s,d] for s in ab_staff] +
               [shuunin_use_a[s,d] for s in shuunin_list])
        model.Add(sum(a_e) == 1)

        a_l = ([x[s,d,"遅"] for s in non_ojt_staff if unit_map.get(s) == "A" and s not in ab_staff_set] +
               [ula[s,d] for s in ab_staff])
        model.Add(sum(a_l) == 1)

        b_e = ([x[s,d,"早"] for s in non_ojt_staff if unit_map.get(s) == "B" and s not in ab_staff_set] +
               [ueb[s,d] for s in ab_staff] +
               [shuunin_use_b[s,d] for s in shuunin_list])
        model.Add(sum(b_e) == 1)

        b_l = ([x[s,d,"遅"] for s in non_ojt_staff if unit_map.get(s) == "B" and s not in ab_staff_set] +
               [ulb[s,d] for s in ab_staff])
        model.Add(sum(b_l) == 1)

        shuunin_night_vars_d = [xs[s,d,"夜"] for s in shuunin_list
                                if requests.get(s,{}).get(days_norm[d])
                                and requests[s][days_norm[d]][0] == "夜"
                                and requests[s][days_norm[d]][1] == "指定"]
        model.Add(sum(x[s,d,"夜"] for s in non_ojt_staff) + sum(shuunin_night_vars_d) == 1)

    for s in staff:
        nt = sum(x[s,d,"夜"] for d in range(N))
        model.Add(nt >= nmin_map[s])
        model.Add(nt <= nmax_map[s])

    # 前半夜勤NG: ○の人は15日以前に夜勤を入れない
    for s in staff:
        if zenhan_ng_map.get(s, "×") == "○":
            for d, dn in enumerate(days_norm):
                if dn.day <= 15:
                    req = requests.get(s, {}).get(dn)
                    if req and req[0] == "夜" and req[1] == "指定":
                        continue
                    model.Add(x[s, d, "夜"] == 0)

    for s in shuunin_list:
        for d in range(N):
            req = requests.get(s, {}).get(days_norm[d])
            if req and req[0] == "夜" and req[1] == "指定":
                continue 
            model.Add(xs[s,d,"夜"] == 0)

    cn_vars = {}
    for s in staff:
        can_consec = (consec_night_map.get(s, "×") == "○")
        for d in range(N - 1):
            if can_consec:
                for sh in ["早","遅","日","×","公","△"]:
                    model.Add(x[s,d+1,sh] == 0).OnlyEnforceIf(x[s,d,"夜"])
                cn = model.NewBoolVar(f"cn_{s}_{d}")
                cn_vars[s,d] = cn
                model.AddBoolAnd([x[s,d,"夜"], x[s,d+1,"夜"]]).OnlyEnforceIf(cn)
                model.AddBoolOr([x[s,d,"夜"].Not(), x[s,d+1,"夜"].Not()]).OnlyEnforceIf(cn.Not())
                if d + 3 < N:
                    model.Add(x[s,d+2,"○"] == 1).OnlyEnforceIf(cn)
                    for sh_w in ["早","遅","日","夜"]:
                        model.Add(x[s,d+3,sh_w] == 0).OnlyEnforceIf(cn)
                elif d + 2 < N:
                    model.Add(x[s,d+2,"○"] == 1).OnlyEnforceIf(cn)
                if d + 2 < N:
                    model.Add(x[s,d,"夜"] + x[s,d+1,"夜"] + x[s,d+2,"夜"] <= 2)
            else:
                for sh_forbidden in ["早","遅","日","夜","×","公","△"]:
                    model.Add(x[s,d+1,sh_forbidden] == 0).OnlyEnforceIf(x[s,d,"夜"])
        
        # 連続夜勤希望者は必ず1度(2連続)夜勤を入れる
        if can_consec:
            model.Add(sum(cn_vars[s,d] for d in range(N-1)) == 1)

    for s in shuunin_list:
        for d in range(N - 1):
            model.Add(xs[s,d+1,"○"] == 1).OnlyEnforceIf(xs[s,d,"夜"])

    for s in staff:
        for d in range(N):
            if d == 0:
                prev_seq = prev_month.get(s, [])
                if not (prev_seq and prev_seq[-1] == "夜"):
                    model.Add(x[s, 0, "○"] == 0)
            else:
                model.Add(x[s, d, "○"] == 0).OnlyEnforceIf(x[s, d-1, "夜"].Not())
                # △(standalone日) は夜勤明けには使えない
                model.Add(x[s, d, "△"] == 0).OnlyEnforceIf(x[s, d-1, "夜"])

    for s in shuunin_list:
        for d in range(N):
            if d == 0:
                prev_seq = prev_month.get(s, [])
                if not (prev_seq and prev_seq[-1] == "夜"):
                    model.Add(xs[s, 0, "○"] == 0)
            else:
                model.Add(xs[s, d, "○"] == 0).OnlyEnforceIf(xs[s, d-1, "夜"].Not())
                model.Add(xs[s, d, "△"] == 0).OnlyEnforceIf(xs[s, d-1, "夜"])

    for s in staff:
        for d in range(N - 1):
            model.Add(x[s,d,"遅"] + x[s,d+1,"早"] <= 1)
    for s in shuunin_list:
        for d in range(N - 1):
            model.Add(xs[s,d,"遅"] + xs[s,d+1,"早"] <= 1)

    for s in staff:
        for date_obj, (sh_type, req_type) in requests.get(s, {}).items():
            if req_type == "希望" and sh_type in ["×","有","公","△"]:
                for d, dn in enumerate(days_norm):
                    if dn == date_obj and d > 0:
                        model.Add(x[s,d-1,"夜"] == 0)
                        break

    for s in all_staff_names:
        max_c  = 5 if cont_map[s] == "40h" else 4
        prev_c = count_trailing_consec(prev_month.get(s, []))
        remain = max(0, max_c - prev_c)
        var_d = xs if s in shuunin_list else x
        if prev_c > 0 and remain < max_c:
            for w in range(1, min(remain + 2, N + 1)):
                if w > remain:
                    model.Add(sum(var_d[s,d2,sh2] for d2 in range(w)
                                  for sh2 in ["早","遅","夜","日","有"]) <= remain)
                    break
        for st in range(N - max_c):
            model.Add(sum(var_d[s,d2,sh2] for d2 in range(st, st+max_c+1)
                          for sh2 in ["早","遅","夜","日","有"]) <= max_c)

    for s in all_staff_names:
        var_d = xs if s in shuunin_list else x
        for sh in ["早", "遅", "日"]:
            prev_seq = prev_month.get(s, [])
            prev_sh_c = 0
            for ps in reversed(prev_seq):
                if ps == sh: prev_sh_c += 1
                else: break
            
            if prev_sh_c >= 2:
                model.Add(var_d[s, 0, sh] == 0)
            elif prev_sh_c == 1:
                if N >= 2:
                    model.Add(var_d[s, 0, sh] + var_d[s, 1, sh] <= 1)
            
            for d in range(N - 2):
                model.Add(var_d[s, d, sh] + var_d[s, d+1, sh] + var_d[s, d+2, sh] <= 2)

    # 主任には休みの下限のみ設定（より多くの休み=×を取れるようにする）
    for s in all_staff_names:
        min_hol = holiday_limits.get(cont_map[s], 0)
        var_d = xs if s in shuunin_list else x
        total_off = (sum(var_d[s,d,"×"] for d in range(N)) +
                     sum(var_d[s,d,"○"] for d in range(N)) +
                     sum(var_d[s,d,"公"] for d in range(N)) +
                     sum(var_d[s,d,"△"] for d in range(N)))
        if s in shuunin_list:
            model.Add(total_off >= min_hol)
        else:
            model.Add(total_off == min_hol)

    for s in all_staff_names:
        var_d = xs if s in shuunin_list else x
        allowed = allowed_shifts_map.get(s)
        if allowed is not None:
            forbidden = set(WORK_SHIFTS) - allowed
            for d in range(N):
                for sh in forbidden:
                    req = requests.get(s, {}).get(days_norm[d])
                    if req and req[1] == "指定" and req[0] == sh: continue
                    model.Add(var_d[s,d,sh] == 0)
        
        if s in weekday_allowed_map:
            for d in range(N):
                wd = days_norm[d].weekday()
                if wd in weekday_allowed_map[s]:
                    allowed_wd = weekday_allowed_map[s][wd]
                    all_possible = set(WORK_SHIFTS) | {"×"}
                    forbidden_wd = all_possible - allowed_wd
                    for sh in forbidden_wd:
                        req = requests.get(s, {}).get(days_norm[d])
                        if req and req[1] == "指定" and req[0] == sh: continue
                        if sh == "×":
                            model.Add(sum(var_d[s,d,sh2] for sh2 in WORK_SHIFTS) == 1)
                        else:
                            model.Add(var_d[s,d,sh] == 0)

    for s in part_staff:
        nen_limit = nenkyuu_limits.get(cont_map.get(s, "40h"), 0)
        if nen_limit > 0:
            continue
        for d in range(N):
            req = requests.get(s, {}).get(days_norm[d])
            if req and req[0] == "有" and req[1] == "指定":
                pass
            else:
                model.Add(x[s,d,"有"] == 0)

    # 「公」は 32h スタッフのみ使用可能（40h・パートは禁止）
    for s in all_staff_names:
        var_d = xs if s in shuunin_list else x
        if cont_map.get(s, "40h") != "32h":
            for d in range(N):
                req = requests.get(s, {}).get(days_norm[d])
                if req and req[1] == "指定" and req[0] == "公":
                    continue
                model.Add(var_d[s, d, "公"] == 0)

    for s in all_staff_names:
        nen_limit = nenkyuu_limits.get(cont_map.get(s, "40h"), 2)
        var_d = xs if s in shuunin_list else x
        total_nenkyuu = sum(var_d[s,d,"有"] for d in range(N))
        if nen_limit > 0:
            model.Add(total_nenkyuu == nen_limit)
        else:
            model.Add(total_nenkyuu == 0)

    for s in staff:
        if s not in weekly_work_days:
            continue
        target = weekly_work_days[s]
        for week_key in sorted_week_keys:
            didx = week_groups[week_key]
            wv = [x[s,d,sh] for d in didx for sh in ["早","遅","夜","有","日"]]
            if len(didx) == 7:
                model.Add(sum(wv) >= max(0, target - 1))
                model.Add(sum(wv) <= target)
            else:
                model.Add(sum(wv) <= round(target * len(didx) / 7 + 0.5))

    # 主任は「日勤」「△（standalone日休み）」「公」を担当しない（指定がある場合のみ例外）
    for s in shuunin_list:
        for d in range(N):
            req = requests.get(s, {}).get(days_norm[d])
            for sh in ["遅","夜","日","△","公"]:
                if sh in ["遅","夜"] and req and req[0] == sh and req[1] == "指定":
                    continue
                model.Add(xs[s,d,sh] == 0)

    for wd_target in nikkin_days_settings:
        for d in range(N):
            if days_norm[d].weekday() == wd_target:
                model.Add(sum(x[s,d,"日"] for s in staff) >= 1)

    penalty_terms = []

    # ── OJT 制約 ─────────────────────────────────────────────────────
    for s in ojt_list:
        instr = ojt_instructor.get(s)
        if not instr or instr not in all_staff_names:
            continue
        instr_vars = xs if instr in shuunin_list else x

        for d in range(N):
            # ハード制約: 指導者が夜勤の日はOJTを夜勤以外にする
            if instr not in shuunin_list:
                model.Add(x[s, d, "夜"] == 0).OnlyEnforceIf(x[instr, d, "夜"])
            else:
                # 主任が指導者の場合（稀）も同様
                model.Add(x[s, d, "夜"] == 0).OnlyEnforceIf(xs[instr, d, "夜"])

            # ソフト制約: 指導者と同じシフトを入れる（ペナルティ25）
            for sh in ALL_SHIFTS:
                diff = model.NewBoolVar(f"ojt_diff_{s}_{d}_{sh}")
                model.Add(x[s, d, sh] != instr_vars[instr, d, sh]).OnlyEnforceIf(diff)
                model.Add(x[s, d, sh] == instr_vars[instr, d, sh]).OnlyEnforceIf(diff.Not())
                penalty_terms.append((diff, 25))

    # 期間別公休数制約（土/日/公 それぞれ個別に強制）
    # 土 = × のみ、日 = ○(夜勤明け) + △(standalone日)、公 = 公
    for s in all_staff_names:
        var_d = xs if s in shuunin_list else x
        s_type = cont_map.get(s, "40h")
        for (p_start, p_end, p_type, n_土, n_日, n_公) in holiday_periods:
            if p_type != s_type:
                continue
            period_d_indices = [d for d, dn in enumerate(days_norm) if p_start <= dn <= p_end]
            if not period_d_indices:
                continue
            period_土 = sum(var_d[s,d,"×"] for d in period_d_indices)
            period_日 = (sum(var_d[s,d,"○"] for d in period_d_indices) +
                         sum(var_d[s,d,"△"] for d in period_d_indices))
            period_公 = sum(var_d[s,d,"公"] for d in period_d_indices)
            period_off = period_土 + period_日 + period_公
            p_count = n_土 + n_日 + n_公
            if s in shuunin_list:
                model.Add(period_off >= p_count)
            else:
                # ハード下限 + 超過をペナルティ化
                if n_土 > 0:
                    model.Add(period_土 >= n_土)
                    diff_土 = model.NewIntVar(0, N, f"m1_diff_土_{s}_{p_type}_{p_start.day}")
                    model.Add(diff_土 >= period_土 - n_土)
                    penalty_terms.append((diff_土, 500))
                else:
                    diff_土0 = model.NewIntVar(0, N, f"m1_diff_土0_{s}_{p_type}_{p_start.day}")
                    model.Add(diff_土0 == period_土)
                    penalty_terms.append((diff_土0, 500))
                if n_日 > 0:
                    model.Add(period_日 >= n_日)
                    diff_日 = model.NewIntVar(0, N, f"m1_diff_日_{s}_{p_type}_{p_start.day}")
                    model.Add(diff_日 >= period_日 - n_日)
                    penalty_terms.append((diff_日, 500))
                else:
                    diff_日0 = model.NewIntVar(0, N, f"m1_diff_日0_{s}_{p_type}_{p_start.day}")
                    model.Add(diff_日0 == period_日)
                    penalty_terms.append((diff_日0, 500))
                if n_公 > 0:
                    model.Add(period_公 >= n_公)
                    diff_公 = model.NewIntVar(0, N, f"m1_diff_公_{s}_{p_type}_{p_start.day}")
                    model.Add(diff_公 >= period_公 - n_公)
                    penalty_terms.append((diff_公, 500))
                else:
                    diff_公0 = model.NewIntVar(0, N, f"m1_diff_公0_{s}_{p_type}_{p_start.day}")
                    model.Add(diff_公0 == period_公)
                    penalty_terms.append((diff_公0, 500))
                diff_tot = model.NewIntVar(0, N, f"m1_diff_tot_{s}_{p_type}_{p_start.day}")
                model.Add(diff_tot >= period_off - p_count)
                penalty_terms.append((diff_tot, 10))

    # 主任使用ペナルティ（究極の最終手段として極大設定）
    for s in shuunin_list:
        for d in range(N):
            penalty_terms.append((xs[s,d,"早"], 100000))

    # その他ソフト制約群
    for s in staff:
        avg_night = (nmin_map[s] + nmax_map[s]) / 2.0
        target_n = int(avg_night + 0.5)
        actual_n = model.NewIntVar(0, N, f"actual_n_{s}")
        model.Add(actual_n == sum(x[s, d, "夜"] for d in range(N)))
        diff_n = model.NewIntVar(0, N, f"diff_n_{s}")
        model.Add(diff_n >= actual_n - target_n)
        model.Add(diff_n >= target_n - actual_n)
        penalty_terms.append((diff_n, 15))

    non_leader = [s for s in staff if role_map.get(s) != "リーダー"]
    if len(non_leader) >= 2:
        e_vars = []; l_vars = []
        for s in non_leader:
            ev = model.NewIntVar(0, N, f"e_{s}")
            lv = model.NewIntVar(0, N, f"l_{s}")
            model.Add(ev == sum(x[s,d,"早"] for d in range(N)))
            model.Add(lv == sum(x[s,d,"遅"] for d in range(N)))
            e_vars.append(ev); l_vars.append(lv)
        max_e = model.NewIntVar(0, N, "max_e"); min_e = model.NewIntVar(0, N, "min_e")
        max_l = model.NewIntVar(0, N, "max_l"); min_l = model.NewIntVar(0, N, "min_l")
        model.AddMaxEquality(max_e, e_vars); model.AddMinEquality(min_e, e_vars)
        model.AddMaxEquality(max_l, l_vars); model.AddMinEquality(min_l, l_vars)
        diff_e = model.NewIntVar(0, N, "diff_e"); model.Add(diff_e == max_e - min_e)
        diff_l = model.NewIntVar(0, N, "diff_l"); model.Add(diff_l == max_l - min_l)
        penalty_terms.append((diff_e, 5))
        penalty_terms.append((diff_l, 5))

    for s in staff:
        for d in range(N - 1):
            late_then_day = model.NewBoolVar(f"ltd_{s}_{d}")
            model.AddBoolAnd([x[s,d,"遅"], x[s,d+1,"日"]]).OnlyEnforceIf(late_then_day)
            model.AddBoolOr([x[s,d,"遅"].Not(), x[s,d+1,"日"].Not()]).OnlyEnforceIf(late_then_day.Not())
            penalty_terms.append((late_then_day, 10))

    for s in staff:
        for start in range(N - 10):
            rest_bits = ([x[s,d,"×"] for d in range(start, start + 11)] +
                         [x[s,d,"公"] for d in range(start, start + 11)] +
                         [x[s,d,"△"] for d in range(start, start + 11)])
            gap_viol = model.NewBoolVar(f"gv_{s}_{start}")
            model.Add(sum(rest_bits) == 0).OnlyEnforceIf(gap_viol)
            model.Add(sum(rest_bits) >= 1).OnlyEnforceIf(gap_viol.Not())
            penalty_terms.append((gap_viol, 50))

    for s in staff:
        if s in part_with_fixed:
            continue
        for d in range(N - 3):
            work_d = [model.NewBoolVar(f"wd4_{s}_{d}_{k}") for k in range(4)]
            for k in range(4):
                model.Add(sum(x[s,d+k,sh] for sh in ["早","遅","夜","日","有"]) == 1
                          ).OnlyEnforceIf(work_d[k])
                model.Add(sum(x[s,d+k,sh] for sh in ["早","遅","夜","日","有"]) == 0
                          ).OnlyEnforceIf(work_d[k].Not())
            w4_real = model.NewBoolVar(f"w4r_{s}_{d}")
            model.AddBoolAnd(work_d).OnlyEnforceIf(w4_real)
            model.AddBoolOr([w.Not() for w in work_d]).OnlyEnforceIf(w4_real.Not())
            penalty_terms.append((w4_real, 2))

    for s in part_staff:
        for d in range(N - 3):
            work_d_p = [model.NewBoolVar(f"wd4p_{s}_{d}_{k}") for k in range(4)]
            for k in range(4):
                model.Add(sum(x[s,d+k,sh] for sh in ["早","遅","夜","日","有"]) == 1).OnlyEnforceIf(work_d_p[k])
                model.Add(sum(x[s,d+k,sh] for sh in ["早","遅","夜","日","有"]) == 0).OnlyEnforceIf(work_d_p[k].Not())
            w4_p = model.NewBoolVar(f"w4p_{s}_{d}")
            model.AddBoolAnd(work_d_p).OnlyEnforceIf(w4_p)
            model.AddBoolOr([w.Not() for w in work_d_p]).OnlyEnforceIf(w4_p.Not())
            penalty_terms.append((w4_p, 50))

    obj_terms = []
    for var, coef in penalty_terms:
        obj_terms.append(var * coef)
    if obj_terms:
        model.Minimize(sum(obj_terms))

    # validate_only モード: 短時間で実現可能性だけを確認
    if validate_only:
        solver_v = cp_model.CpSolver()
        solver_v.parameters.max_time_in_seconds = 15
        solver_v.parameters.num_search_workers  = num_workers
        solver_v.parameters.random_seed         = 0
        # 目的関数なし（実現可能解を探すだけ）
        model_v = cp_model.CpModel()
        # 同じ制約をそのまま使う（model は既に構築済み）
        # → model をそのまま使ってソルブ（Minimize は無視されない問題があるので
        #    タイムアウト内に FEASIBLE が出れば OK）
        status_v = solver_v.Solve(model)
        if status_v in (cp_model.FEASIBLE, cp_model.OPTIMAL):
            return {"feasible": True, "messages": ["✓ 制約チェック通過: 実現可能な解が存在します。計算を開始できます。"]}
        else:
            # INFEASIBLE or timeout → 詳細診断
            diag_msgs = _diagnose_infeasible(
                staff, shuunin_list, requests, days_norm, N,
                allowed_shifts_map, fixed_holiday_map, holiday_limits,
                cont_map, nmin_map, nmax_map, prev_month, weekly_work_days,
                unit_map=unit_map, ab_staff_set=ab_staff_set,
                weekday_allowed_map=weekday_allowed_map,
                nikkin_days_settings=nikkin_days_settings,
                ojt_list=ojt_list
            )
            if status_v == cp_model.INFEASIBLE:
                prefix_msg = "制約が矛盾しているため解が存在しません。"
            else:
                prefix_msg = "15秒以内に実現可能解が見つかりませんでした（制約が厳しすぎる可能性があります）。"
            all_msgs = [prefix_msg] + (diag_msgs or ["詳細な原因を特定できませんでした。"])
            # 緩和提案を追加
            suggestions = _suggest_relaxation("\n".join(all_msgs))
            if suggestions:
                all_msgs += ["", "【緩和提案】"] + suggestions
            return {"feasible": False, "messages": all_msgs}

    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = timeout
    solver.parameters.num_search_workers  = num_workers
    solver.parameters.random_seed         = random_seed
    _push_progress(progress_uid, "モデル1: ソルバー開始...")
    cb1 = ProgressCallback(progress_uid, N, "M1")
    status = solver.Solve(model, cb1)

    if status not in (cp_model.FEASIBLE, cp_model.OPTIMAL):
        model2 = cp_model.CpModel()
        penalty2 = []

        x2 = {}
        for s in staff:
            for d in range(N):
                for sh in ALL_SHIFTS:
                    x2[s, d, sh] = model2.NewBoolVar(f"x2_{s}_{d}_{sh}")
        xs2 = {}
        for s in shuunin_list:
            for d in range(N):
                for sh in ALL_SHIFTS:
                    xs2[s, d, sh] = model2.NewBoolVar(f"xs2_{s}_{d}_{sh}")

        uea2 = {}; ueb2 = {}; ula2 = {}; ulb2 = {}
        for s in ab_staff:
            for d in range(N):
                uea2[s,d] = model2.NewBoolVar(f"uea2_{s}_{d}")
                ueb2[s,d] = model2.NewBoolVar(f"ueb2_{s}_{d}")
                ula2[s,d] = model2.NewBoolVar(f"ula2_{s}_{d}")
                ulb2[s,d] = model2.NewBoolVar(f"ulb2_{s}_{d}")
                model2.Add(uea2[s,d] + ueb2[s,d] == x2[s,d,"早"])
                model2.Add(ula2[s,d] + ulb2[s,d] == x2[s,d,"遅"])

        shuunin_use_a2 = {}; shuunin_use_b2 = {}
        for s in shuunin_list:
            for d in range(N):
                shuunin_use_a2[s,d] = model2.NewBoolVar(f"sh_ua2_{s}_{d}")
                shuunin_use_b2[s,d] = model2.NewBoolVar(f"sh_ub2_{s}_{d}")
                model2.Add(shuunin_use_a2[s,d] + shuunin_use_b2[s,d] <= xs2[s,d,"早"])
                model2.Add(shuunin_use_a2[s,d] + shuunin_use_b2[s,d] <= 1)

        def _rebuild_model2():
            nonlocal cn2_vars
            for s in staff:
                for d in range(N):
                    model2.AddExactlyOne(x2[s,d,sh] for sh in ALL_SHIFTS)
            for s in shuunin_list:
                for d in range(N):
                    model2.AddExactlyOne(xs2[s,d,sh] for sh in ALL_SHIFTS)

            for s in staff:
                if s not in requests: continue
                for date_obj, (sh_type, req_type) in requests[s].items():
                    for d, dn in enumerate(days_norm):
                        if dn == date_obj and sh_type in ALL_SHIFTS:
                            model2.Add(x2[s,d,sh_type] == 1)
                            break
            for s in shuunin_list:
                if s not in requests: continue
                for date_obj, (sh_type, req_type) in requests[s].items():
                    for d, dn in enumerate(days_norm):
                        if dn == date_obj and sh_type in ALL_SHIFTS:
                            model2.Add(xs2[s,d,sh_type] == 1)
                            break

            for s in staff:
                if prev_month.get(s, []) and prev_month[s][-1] == "夜":
                    for sh_f in ["早","遅","日","夜","×"]:
                        model2.Add(x2[s,0,sh_f] == 0)

            for s, wdays in fixed_holiday_map.items():
                var_dict = xs2 if s in shuunin_list else x2
                for d_idx, dn in enumerate(days_norm):
                    if dn.weekday() in wdays:
                        req = requests.get(s, {}).get(dn)
                        if req and req[1] == "指定": continue
                        model2.Add(var_dict[s,d_idx,"×"] == 1)

            for d in range(N):
                a_e2 = ([x2[s,d,"早"] for s in non_ojt_staff if unit_map.get(s) == "A" and s not in ab_staff_set] +
                        [uea2[s,d] for s in ab_staff] +
                        [shuunin_use_a2[s,d] for s in shuunin_list])
                model2.Add(sum(a_e2) == 1)
                a_l2 = ([x2[s,d,"遅"] for s in non_ojt_staff if unit_map.get(s) == "A" and s not in ab_staff_set] +
                        [ula2[s,d] for s in ab_staff])
                model2.Add(sum(a_l2) == 1)
                b_e2 = ([x2[s,d,"早"] for s in non_ojt_staff if unit_map.get(s) == "B" and s not in ab_staff_set] +
                        [ueb2[s,d] for s in ab_staff] +
                        [shuunin_use_b2[s,d] for s in shuunin_list])
                model2.Add(sum(b_e2) == 1)
                b_l2 = ([x2[s,d,"遅"] for s in non_ojt_staff if unit_map.get(s) == "B" and s not in ab_staff_set] +
                        [ulb2[s,d] for s in ab_staff])
                model2.Add(sum(b_l2) == 1)
                shuunin_night_vars_d2 = [xs2[s,d,"夜"] for s in shuunin_list
                                         if requests.get(s,{}).get(days_norm[d])
                                         and requests[s][days_norm[d]][0] == "夜"
                                         and requests[s][days_norm[d]][1] == "指定"]
                model2.Add(sum(x2[s,d,"夜"] for s in non_ojt_staff) + sum(shuunin_night_vars_d2) == 1)

            for s in staff:
                nt2 = sum(x2[s,d,"夜"] for d in range(N))
                model2.Add(nt2 >= nmin_map[s])
                model2.Add(nt2 <= nmax_map[s])

            # 前半夜勤NG（model2）
            for s in staff:
                if zenhan_ng_map.get(s, "×") == "○":
                    for d, dn in enumerate(days_norm):
                        if dn.day <= 15:
                            req = requests.get(s, {}).get(dn)
                            if req and req[0] == "夜" and req[1] == "指定":
                                continue
                            model2.Add(x2[s, d, "夜"] == 0)
            for s in shuunin_list:
                for d in range(N):
                    req = requests.get(s, {}).get(days_norm[d])
                    if req and req[0] == "夜" and req[1] == "指定": continue
                    model2.Add(xs2[s,d,"夜"] == 0)

            cn2_vars = {}
            for s in staff:
                can_consec = (consec_night_map.get(s, "×") == "○")
                for d in range(N - 1):
                    if can_consec:
                        for sh in ["早","遅","日","×","公","△"]:
                            model2.Add(x2[s,d+1,sh] == 0).OnlyEnforceIf(x2[s,d,"夜"])
                        cn2 = model2.NewBoolVar(f"cn2_{s}_{d}")
                        cn2_vars[s,d] = cn2
                        model2.AddBoolAnd([x2[s,d,"夜"], x2[s,d+1,"夜"]]).OnlyEnforceIf(cn2)
                        model2.AddBoolOr([x2[s,d,"夜"].Not(), x2[s,d+1,"夜"].Not()]).OnlyEnforceIf(cn2.Not())
                        if d + 3 < N:
                            model2.Add(x2[s,d+2,"○"] == 1).OnlyEnforceIf(cn2)
                            for sh_w in ["日","有"]:
                                model2.Add(x2[s,d+3,sh_w] == 0).OnlyEnforceIf(cn2)
                        elif d + 2 < N:
                            model2.Add(x2[s,d+2,"○"] == 1).OnlyEnforceIf(cn2)
                        if d + 2 < N:
                            model2.Add(x2[s,d,"夜"] + x2[s,d+1,"夜"] + x2[s,d+2,"夜"] <= 2)
                    else:
                        for sh_forbidden in ["早","遅","日","夜","×","公","△"]:
                            model2.Add(x2[s,d+1,sh_forbidden] == 0).OnlyEnforceIf(x2[s,d,"夜"])
                if can_consec:
                    model2.Add(sum(cn2_vars[s,d] for d in range(N-1)) == 1)

            for s in shuunin_list:
                for d in range(N - 1):
                    model2.Add(xs2[s,d+1,"○"] == 1).OnlyEnforceIf(xs2[s,d,"夜"])

            for s in staff:
                for d in range(N):
                    if d == 0:
                        prev_seq = prev_month.get(s, [])
                        if not (prev_seq and prev_seq[-1] == "夜"):
                            model2.Add(x2[s, 0, "○"] == 0)
                    else:
                        model2.Add(x2[s, d, "○"] == 0).OnlyEnforceIf(x2[s, d-1, "夜"].Not())
                        model2.Add(x2[s, d, "△"] == 0).OnlyEnforceIf(x2[s, d-1, "夜"])
            for s in shuunin_list:
                for d in range(N):
                    if d == 0:
                        prev_seq = prev_month.get(s, [])
                        if not (prev_seq and prev_seq[-1] == "夜"):
                            model2.Add(xs2[s, 0, "○"] == 0)
                    else:
                        model2.Add(xs2[s, d, "○"] == 0).OnlyEnforceIf(xs2[s, d-1, "夜"].Not())
                        model2.Add(xs2[s, d, "△"] == 0).OnlyEnforceIf(xs2[s, d-1, "夜"])

            for s in staff:
                for d in range(N - 1):
                    model2.Add(x2[s,d,"遅"] + x2[s,d+1,"早"] <= 1)
            for s in shuunin_list:
                for d in range(N - 1):
                    model2.Add(xs2[s,d,"遅"] + xs2[s,d+1,"早"] <= 1)

            for s in staff:
                for date_obj, (sh_type, req_type) in requests.get(s, {}).items():
                    if req_type == "希望" and sh_type in ["×","有","公","△"]:
                        for d, dn in enumerate(days_norm):
                            if dn == date_obj and d > 0:
                                model2.Add(x2[s,d-1,"夜"] == 0)
                                break

            for s in all_staff_names:
                max_c  = 5 if cont_map[s] == "40h" else 4
                prev_c = count_trailing_consec(prev_month.get(s, []))
                remain = max(0, max_c - prev_c)
                var_d2 = xs2 if s in shuunin_list else x2
                if prev_c > 0 and remain < max_c:
                    for w in range(1, min(remain + 2, N + 1)):
                        if w > remain:
                            model2.Add(sum(var_d2[s,d2,sh2] for d2 in range(w)
                                          for sh2 in ["早","遅","夜","日","有"]) <= remain)
                            break
                for st in range(N - max_c):
                    model2.Add(sum(var_d2[s,d2,sh2] for d2 in range(st, st+max_c+1)
                                  for sh2 in ["早","遅","夜","日","有"]) <= max_c)

            for s in all_staff_names:
                var_d2 = xs2 if s in shuunin_list else x2
                for sh in ["早", "遅", "日"]:
                    for d in range(N - 2):
                        model2.Add(var_d2[s, d, sh] + var_d2[s, d+1, sh] + var_d2[s, d+2, sh] <= 2)

            for s in all_staff_names:
                min_hol = holiday_limits.get(cont_map[s], 0)
                var_d2 = xs2 if s in shuunin_list else x2
                total_off2 = (sum(var_d2[s,d,"×"] for d in range(N)) +
                              sum(var_d2[s,d,"○"] for d in range(N)) +
                              sum(var_d2[s,d,"公"] for d in range(N)) +
                              sum(var_d2[s,d,"△"] for d in range(N)))
                model2.Add(total_off2 >= min_hol)
                if s not in shuunin_list:
                    diff_hol = model2.NewIntVar(0, N, f"diff_hol_{s}")
                    model2.Add(diff_hol >= total_off2 - min_hol)
                    penalty2.append((diff_hol, 10))

            # 期間別公休数制約（土/日/公 それぞれ個別に、model2）
            # 土 = × のみ、日 = ○(夜勤明け) + △(standalone日)
            for s in all_staff_names:
                var_d2 = xs2 if s in shuunin_list else x2
                s_type = cont_map.get(s, "40h")
                for (p_start, p_end, p_type, n_土, n_日, n_公) in holiday_periods:
                    if p_type != s_type:
                        continue
                    period_d_indices2 = [d for d, dn in enumerate(days_norm) if p_start <= dn <= p_end]
                    if not period_d_indices2:
                        continue
                    p_土2 = sum(var_d2[s,d,"×"] for d in period_d_indices2)
                    p_日2 = (sum(var_d2[s,d,"○"] for d in period_d_indices2) +
                              sum(var_d2[s,d,"△"] for d in period_d_indices2))
                    p_公2 = sum(var_d2[s,d,"公"] for d in period_d_indices2)
                    period_off2 = p_土2 + p_日2 + p_公2
                    p_count = n_土 + n_日 + n_公
                    # 合計の下限
                    model2.Add(period_off2 >= p_count)
                    if s not in shuunin_list:
                        # 各種別ごとに下限 + 上限ペナルティ（model1が失敗した場合も種別を守る）
                        if n_土 > 0:
                            model2.Add(p_土2 >= n_土)
                            diff_土 = model2.NewIntVar(0, N, f"diff_土_{s}_{p_type}_{p_start.day}")
                            model2.Add(diff_土 >= p_土2 - n_土)
                            penalty2.append((diff_土, 500))
                        else:
                            # n_土=0のとき 土を使わせない（高ペナルティ）
                            diff_土0 = model2.NewIntVar(0, N, f"diff_土0_{s}_{p_type}_{p_start.day}")
                            model2.Add(diff_土0 == p_土2)
                            penalty2.append((diff_土0, 500))
                        if n_日 > 0:
                            model2.Add(p_日2 >= n_日)
                            diff_日 = model2.NewIntVar(0, N, f"diff_日_{s}_{p_type}_{p_start.day}")
                            model2.Add(diff_日 >= p_日2 - n_日)
                            penalty2.append((diff_日, 500))
                        else:
                            diff_日0 = model2.NewIntVar(0, N, f"diff_日0_{s}_{p_type}_{p_start.day}")
                            model2.Add(diff_日0 == p_日2)
                            penalty2.append((diff_日0, 500))
                        if n_公 > 0:
                            model2.Add(p_公2 >= n_公)
                            diff_公 = model2.NewIntVar(0, N, f"diff_公_{s}_{p_type}_{p_start.day}")
                            model2.Add(diff_公 >= p_公2 - n_公)
                            penalty2.append((diff_公, 500))
                        else:
                            diff_公0 = model2.NewIntVar(0, N, f"diff_公0_{s}_{p_type}_{p_start.day}")
                            model2.Add(diff_公0 == p_公2)
                            penalty2.append((diff_公0, 500))
                        # 合計超過ペナルティ
                        diff_period = model2.NewIntVar(0, N, f"diff_period_{s}_{p_type}_{p_start.day}")
                        model2.Add(diff_period >= period_off2 - p_count)
                        penalty2.append((diff_period, 10))

            for s in all_staff_names:
                var_d2 = xs2 if s in shuunin_list else x2
                allowed = allowed_shifts_map.get(s)
                if allowed is not None:
                    forbidden = set(WORK_SHIFTS) - allowed
                    for d in range(N):
                        for sh in forbidden:
                            req = requests.get(s, {}).get(days_norm[d])
                            if req and req[1] == "指定" and req[0] == sh: continue
                            model2.Add(var_d2[s,d,sh] == 0)
                
                if s in weekday_allowed_map:
                    for d in range(N):
                        wd = days_norm[d].weekday()
                        if wd in weekday_allowed_map[s]:
                            allowed_wd = weekday_allowed_map[s][wd]
                            all_possible = set(WORK_SHIFTS) | {"×"}
                            forbidden_wd = all_possible - allowed_wd
                            for sh in forbidden_wd:
                                req = requests.get(s, {}).get(days_norm[d])
                                if req and req[1] == "指定" and req[0] == sh: continue
                                if sh == "×":
                                    model2.Add(sum(var_d2[s,d,sh2] for sh2 in WORK_SHIFTS) == 1)
                                else:
                                    model2.Add(var_d2[s,d,sh] == 0)

            for s in part_staff:
                nen_limit = nenkyuu_limits.get(cont_map.get(s, "40h"), 0)
                if nen_limit > 0: continue
                for d in range(N):
                    req = requests.get(s, {}).get(days_norm[d])
                    if req and req[0] == "有" and req[1] == "指定": pass
                    else: model2.Add(x2[s,d,"有"] == 0)

            # 「公」は 32h スタッフのみ使用可能（model2）
            for s in all_staff_names:
                var_d2 = xs2 if s in shuunin_list else x2
                if cont_map.get(s, "40h") != "32h":
                    for d in range(N):
                        req = requests.get(s, {}).get(days_norm[d])
                        if req and req[1] == "指定" and req[0] == "公":
                            continue
                        model2.Add(var_d2[s, d, "公"] == 0)

            for s in all_staff_names:
                if s in part_staff: continue
                nen_limit = nenkyuu_limits.get(cont_map.get(s, "40h"), 2)
                var_d2 = xs2 if s in shuunin_list else x2
                total_nenkyuu2 = sum(var_d2[s,d,"有"] for d in range(N))
                if nen_limit > 0:
                    model2.Add(total_nenkyuu2 == nen_limit)
                else:
                    model2.Add(total_nenkyuu2 == 0)

            for s in staff:
                if s not in weekly_work_days: continue
                target = weekly_work_days[s]
                for week_key in sorted_week_keys:
                    didx = week_groups[week_key]
                    wv2 = [x2[s,d,sh] for d in didx for sh in ["早","遅","夜","有","日"]]
                    if len(didx) == 7:
                        model2.Add(sum(wv2) >= max(0, target - 1))
                        model2.Add(sum(wv2) <= target)
                    else:
                        model2.Add(sum(wv2) <= round(target * len(didx) / 7 + 0.5))

            for s in shuunin_list:
                for d in range(N):
                    req = requests.get(s, {}).get(days_norm[d])
                    for sh in ["遅","夜","日","△","公"]:
                        if sh in ["遅","夜"] and req and req[0] == sh and req[1] == "指定": continue
                        model2.Add(xs2[s,d,sh] == 0)

            for wd_target in nikkin_days_settings:
                for d in range(N):
                    if days_norm[d].weekday() == wd_target:
                        model2.Add(sum(x2[s,d,"日"] for s in staff) >= 1)

        cn2_vars = {}
        _rebuild_model2()

        for s in shuunin_list:
            for d in range(N):
                penalty2.append((xs2[s,d,"早"], 100000))

        # OJT 制約（model2）
        for s in ojt_list:
            instr = ojt_instructor.get(s)
            if not instr or instr not in all_staff_names:
                continue
            instr_vars2 = xs2 if instr in shuunin_list else x2
            for d in range(N):
                model2.Add(x2[s, d, "夜"] == 0).OnlyEnforceIf(instr_vars2[instr, d, "夜"])
                for sh in ALL_SHIFTS:
                    diff2 = model2.NewBoolVar(f"ojt2_diff_{s}_{d}_{sh}")
                    model2.Add(x2[s,d,sh] != instr_vars2[instr,d,sh]).OnlyEnforceIf(diff2)
                    model2.Add(x2[s,d,sh] == instr_vars2[instr,d,sh]).OnlyEnforceIf(diff2.Not())
                    penalty2.append((diff2, 25))
        
        non_leader2 = [s for s in staff if role_map.get(s) != "リーダー"]
        if len(non_leader2) >= 2:
            e2_vars = []; l2_vars = []
            for s in non_leader2:
                ev2 = model2.NewIntVar(0, N, f"e2_{s}")
                lv2 = model2.NewIntVar(0, N, f"l2_{s}")
                model2.Add(ev2 == sum(x2[s,d,"早"] for d in range(N)))
                model2.Add(lv2 == sum(x2[s,d,"遅"] for d in range(N)))
                e2_vars.append(ev2); l2_vars.append(lv2)
            max_e2 = model2.NewIntVar(0, N, "max_e2"); min_e2 = model2.NewIntVar(0, N, "min_e2")
            max_l2 = model2.NewIntVar(0, N, "max_l2"); min_l2 = model2.NewIntVar(0, N, "min_l2")
            model2.AddMaxEquality(max_e2, e2_vars); model2.AddMinEquality(min_e2, e2_vars)
            model2.AddMaxEquality(max_l2, l2_vars); model2.AddMinEquality(min_l2, l2_vars)
            diff_e2 = model2.NewIntVar(0, N, "diff_e2"); model2.Add(diff_e2 == max_e2 - min_e2)
            diff_l2 = model2.NewIntVar(0, N, "diff_l2"); model2.Add(diff_l2 == max_l2 - min_l2)
            penalty2.append((diff_e2, 5))
            penalty2.append((diff_l2, 5))
            
        obj2 = [v * c for v, c in penalty2]
        if obj2:
            model2.Minimize(sum(obj2))

        solver2 = cp_model.CpSolver()
        solver2.parameters.max_time_in_seconds = timeout
        solver2.parameters.num_search_workers  = num_workers
        _push_progress(progress_uid, "モデル1で解が見つからず → モデル2(緩和版)で再試行...")
        cb2 = ProgressCallback(progress_uid, N, "M2")
        status2 = solver2.Solve(model2, cb2)

        if status2 not in (cp_model.FEASIBLE, cp_model.OPTIMAL):
            diag_msgs = _diagnose_infeasible(
                staff, shuunin_list, requests, days_norm, N,
                allowed_shifts_map, fixed_holiday_map, holiday_limits,
                cont_map, nmin_map, nmax_map, prev_month, weekly_work_days,
                unit_map=unit_map, ab_staff_set=ab_staff_set,
                weekday_allowed_map=weekday_allowed_map,
                nikkin_days_settings=nikkin_days_settings,
                ojt_list=ojt_list
            )
            if diag_msgs:
                error_text = "【勤務表を生成できませんでした】\n以下の制約が矛盾している可能性があります：\n\n" + "\n".join(diag_msgs)
                raise Exception(error_text)
            raise Exception(
                "致命的エラー: 主任を使った補充を含めても、条件を満たすシフト表が見つかりませんでした。\n"
                "希望休・指定勤務・夜勤回数・公休数の設定を見直してください。"
            )

        solver = solver2
        x = x2
        xs = xs2
        uea = uea2; ueb = ueb2; ula = ula2; ulb = ulb2
        shuunin_use_a = shuunin_use_a2; shuunin_use_b = shuunin_use_b2

    result = {}
    for s in staff:
        result[s] = {}
        for d in range(N):
            for sh in ALL_SHIFTS:
                if solver.Value(x[s,d,sh]) == 1:
                    result[s][d] = sh
                    break

    for s in shuunin_list:
        result[s] = {}
        for d in range(N):
            for sh in ALL_SHIFTS:
                if solver.Value(xs[s,d,sh]) == 1:
                    result[s][d] = sh
                    break

    ab_unit_result = {}
    for s in ab_staff:
        ab_unit_result[s] = {}
        for d in range(N):
            sh = result[s][d]
            if sh == "早":
                ab_unit_result[s][d] = "A" if solver.Value(uea[s,d]) == 1 else "B"
            elif sh == "遅":
                ab_unit_result[s][d] = "A" if solver.Value(ula[s,d]) == 1 else "B"
            else:
                ab_unit_result[s][d] = None

    shuunin_unit_result = {}
    for s in shuunin_list:
        shuunin_unit_result[s] = {}
        for d in range(N):
            ua = solver.Value(shuunin_use_a[s,d])
            ub = solver.Value(shuunin_use_b[s,d])
            if ua:
                shuunin_unit_result[s][d] = "A"
            elif ub:
                shuunin_unit_result[s][d] = "B"
            else:
                shuunin_unit_result[s][d] = None

    return (result, staff, shuunin_list, unit_map, cont_map, role_map,
            days_norm, requests, ab_unit_result, shuunin_unit_result, kanmu_map, prev_month,
            nmin_map, nmax_map, consec_night_map, holiday_periods,
            ojt_list, ojt_instructor)

# ========================================================
# 自己採点機能
# ========================================================
def score_shift(result, staff, shuunin_list, days_norm, requests,
                prev_month, cont_map, role_map, nmin_map, nmax_map,
                consec_night_map, holiday_periods, unit_map,
                ab_unit_result, shuunin_unit_result, kanmu_map,
                ojt_list=None):
    """生成シフトを自己採点して (total_score, total_max, all_rows) を返す。
    all_rows: [(項目名, 件数, 1件あたり減点, 合計減点, 減点あり?), ...]
              減点がない項目も含む全採点基準を返す
    OJTスタッフは採点対象外。
    """
    _ojt = set(ojt_list or [])
    TOTAL = 1000
    OFF_SHIFTS = {"×", "有", "公", "△", "○"}
    results_dict = {}

    def record(name, count, per):
        pts = int(count * per) if isinstance(count, float) else count * per
        results_dict[name] = (round(count, 1) if isinstance(count, float) else count, per, pts)

    def get_sh(s, d):
        return result.get(s, {}).get(d, "×")

    def is_off(sh):
        return sh in OFF_SHIFTS

    N = len(days_norm)
    # OJTを除いたスタッフリスト（採点対象外）
    score_staff   = [s for s in staff      if s not in _ojt]
    score_members = [s for s in (shuunin_list + staff) if s not in _ojt]

    # ── 1. 夜勤回数が目標値からずれ ──────────────────────────────────
    count = 0
    for s in score_staff:
        target = (nmin_map.get(s, 0) + nmax_map.get(s, 0)) / 2.0
        actual = sum(1 for d in range(N) if get_sh(s, d) == "夜")
        count += abs(actual - target)
    record("①夜勤回数が目標値からずれ", count, 10)

    # ── 2. 早出回数のばらつき（リーダー以外） ────────────────────────
    non_leader = [s for s in score_staff if role_map.get(s) != "リーダー"]
    if len(non_leader) >= 2:
        counts_e = [sum(1 for d in range(N) if get_sh(s, d) == "早") for s in non_leader]
        record("②早出回数のばらつき(max-min)", max(counts_e) - min(counts_e), 5)
    else:
        record("②早出回数のばらつき(max-min)", 0, 5)

    # ── 3. 遅出回数のばらつき（リーダー以外） ────────────────────────
    if len(non_leader) >= 2:
        counts_l = [sum(1 for d in range(N) if get_sh(s, d) == "遅") for s in non_leader]
        record("③遅出回数のばらつき(max-min)", max(counts_l) - min(counts_l), 5)
    else:
        record("③遅出回数のばらつき(max-min)", 0, 5)

    # ── 4. 遅出→翌日早出 ─────────────────────────────────────────────
    count = sum(1 for s in score_members for d in range(N-1)
                if get_sh(s, d) == "遅" and get_sh(s, d+1) == "早")
    record("④遅出→翌日早出", count, 20)

    # ── 5. 11日連続で休みなし ────────────────────────────────────────
    count = sum(1 for s in score_staff for start in range(N-10)
                if all(not is_off(get_sh(s, d)) for d in range(start, start+11)))
    record("⑤11日連続で休みなし", count, 50)

    # ── 6. 4日連続勤務 ───────────────────────────────────────────────
    count = sum(1 for s in score_staff for start in range(N-3)
                if all(not is_off(get_sh(s, d)) for d in range(start, start+4)))
    record("⑥4日連続勤務", count, 5)

    # ── 7. 主任補充使用 ──────────────────────────────────────────────
    count = sum(1 for s in shuunin_list for d in range(N) if get_sh(s, d) == "早")
    record("⑦主任補充使用", count, 80)

    # ── 8. 土/日/公の期間別超過 ──────────────────────────────────────
    count = 0
    for s in score_staff:
        s_type = cont_map.get(s, "40h")
        for (p_start, p_end, p_type, n_土, n_日, n_公) in holiday_periods:
            if p_type != s_type: continue
            didx = [d for d, dn in enumerate(days_norm) if p_start <= dn <= p_end]
            if not didx: continue
            act_土 = sum(1 for d in didx if get_sh(s, d) == "×")
            act_日 = sum(1 for d in didx if get_sh(s, d) in ("○", "△"))
            act_公 = sum(1 for d in didx if get_sh(s, d) == "公")
            count += max(0, act_土 - n_土) + max(0, act_日 - n_日) + max(0, act_公 - n_公)
    record("⑧土/日/公の期間別超過", count, 15)

    # ── 9. 3連続早出 ─────────────────────────────────────────────────
    count = sum(1 for s in score_members for d in range(N-2)
                if get_sh(s, d) == "早" and get_sh(s, d+1) == "早" and get_sh(s, d+2) == "早")
    record("⑨3連続早出", count, 10)

    # ── 10. 3連続遅出 ────────────────────────────────────────────────
    count = sum(1 for s in score_members for d in range(N-2)
                if get_sh(s, d) == "遅" and get_sh(s, d+1) == "遅" and get_sh(s, d+2) == "遅")
    record("⑩3連続遅出", count, 10)

    # ── 12. 週に休みが1日以下 ────────────────────────────────────────
    from collections import defaultdict as _dd
    week_groups = _dd(list)
    for d, dn in enumerate(days_norm):
        sun_offset = (dn.weekday() + 1) % 7
        week_sun = dn - timedelta(days=sun_offset)
        week_groups[week_sun].append(d)
    count = sum(1 for s in score_staff for wk, didx in week_groups.items()
                if len(didx) == 7 and sum(1 for d in didx if is_off(get_sh(s, d))) <= 1)
    record("⑫週の休みが1日以下", count, 20)

    # ── 13. 月末月初で連続勤務 ───────────────────────────────────────
    count = sum(1 for s in score_staff
                if prev_month.get(s, []) and not is_off(prev_month[s][-1])
                and prev_month[s][-1] != "夜" and N > 0 and not is_off(get_sh(s, 0)))
    record("⑬前月末→当月1日が連続勤務", count, 10)

    # ── 14. 夜勤間隔が7日以内（連続夜勤希望○は除く）────────────────
    count = 0
    for s in score_staff:
        if consec_night_map.get(s, "×") == "○": continue
        night_days = [d for d in range(N) if get_sh(s, d) == "夜"]
        for i in range(len(night_days) - 1):
            if night_days[i+1] - night_days[i] <= 7:
                count += 1
    record("⑭夜勤間隔7日以内（連続夜勤希望◯除く）", count, 15)

    # ── 15. 夜勤が前半/後半に偏る（連続夜勤希望○は除く）────────────
    mid = N // 2
    count = 0
    for s in score_staff:
        if consec_night_map.get(s, "×") == "○": continue
        n_first  = sum(1 for d in range(mid)    if get_sh(s, d) == "夜")
        n_second = sum(1 for d in range(mid, N) if get_sh(s, d) == "夜")
        if (n_first + n_second) >= 2 and abs(n_first - n_second) >= 3:
            count += 1
    record("⑮夜勤が前半/後半に偏る（連続夜勤希望◯除く）", count, 10)

    # ── 19. 同契約区分内の総勤務日数ばらつき ─────────────────────────
    for ct in ["40h", "32h"]:
        grp = [s for s in score_staff if cont_map.get(s) == ct]
        if len(grp) >= 2:
            wc = [sum(1 for d in range(N) if not is_off(get_sh(s, d))) for s in grp]
            record(f"⑲総勤務日数ばらつき({ct})", max(wc) - min(wc), 5)
        else:
            record(f"⑲総勤務日数ばらつき({ct})", 0, 5)

    # ── 20. 夜勤回数のばらつき ────────────────────────────────────────
    night_capable = [s for s in score_staff if nmax_map.get(s, 0) > 0]
    if len(night_capable) >= 2:
        nc = [sum(1 for d in range(N) if get_sh(s, d) == "夜") for s in night_capable]
        record("⑳夜勤回数ばらつき", max(nc) - min(nc), 10)
    else:
        record("⑳夜勤回数ばらつき", 0, 10)

    # ── 21. 特定スタッフへの早出集中 ─────────────────────────────────
    if len(non_leader) >= 2:
        counts_e2 = [sum(1 for d in range(N) if get_sh(s, d) == "早") for s in non_leader]
        avg_e = sum(counts_e2) / len(counts_e2)
        over = sum(max(0, c - (avg_e + 3)) for c in counts_e2)
        record("㉑特定スタッフへの早出集中", over, 5)
    else:
        record("㉑特定スタッフへの早出集中", 0, 5)

    # ── 23. 早出が月初/月末に集中 ──────────────────────────────────
    early_days = set(range(3)) | set(range(N-3, N))
    count = sum(1 for s in score_staff
                if sum(1 for d in range(N) if get_sh(s, d) == "早") >= 4
                and sum(1 for d in early_days if get_sh(s, d) == "早") >=
                    sum(1 for d in range(N) if get_sh(s, d) == "早") * 0.5)
    record("㉓早出が月初/月末に集中", count, 5)

    # ── 24. 週4日以上遅出 ────────────────────────────────────────────
    count = sum(1 for s in score_staff for wk, didx in week_groups.items()
                if len(didx) >= 5 and sum(1 for d in didx if get_sh(s, d) == "遅") >= 4)
    record("㉔週4日以上遅出", count, 15)

    # ── 26. 希望勤務が叶わなかった ───────────────────────────────────
    count = 0
    for s in score_members:
        var_d = result.get(s, {})
        for date_obj, (sh_type, req_type) in requests.get(s, {}).items():
            if req_type == "希望" and sh_type in {"早","遅","日","夜"}:
                for d, dn in enumerate(days_norm):
                    if dn == date_obj:
                        if var_d.get(d) != sh_type:
                            count += 1
                        break
    record("㉖希望勤務が叶わなかった", count, 10)

    # ── 27. 前月末遅出→当月1日早出 ───────────────────────────────────
    count = sum(1 for s in score_staff
                if prev_month.get(s, []) and prev_month[s][-1] == "遅"
                and N > 0 and get_sh(s, 0) == "早")
    record("㉗前月末遅出→当月1日早出", count, 15)

    # ── 全項目を順番通りに並べて返す ─────────────────────────────────
    ALL_ITEMS = [
        "①夜勤回数が目標値からずれ",
        "②早出回数のばらつき(max-min)",
        "③遅出回数のばらつき(max-min)",
        "④遅出→翌日早出",
        "⑤11日連続で休みなし",
        "⑥4日連続勤務",
        "⑦主任補充使用",
        "⑧土/日/公の期間別超過",
        "⑨3連続早出",
        "⑩3連続遅出",
        "⑫週の休みが1日以下",
        "⑬前月末→当月1日が連続勤務",
        "⑭夜勤間隔7日以内（連続夜勤希望◯除く）",
        "⑮夜勤が前半/後半に偏る（連続夜勤希望◯除く）",
        "⑲総勤務日数ばらつき(40h)",
        "⑲総勤務日数ばらつき(32h)",
        "⑳夜勤回数ばらつき",
        "㉑特定スタッフへの早出集中",
        "㉓早出が月初/月末に集中",
        "㉔週4日以上遅出",
        "㉖希望勤務が叶わなかった",
        "㉗前月末遅出→当月1日早出",
    ]

    all_rows = []
    for name in ALL_ITEMS:
        if name in results_dict:
            cnt, per, pts = results_dict[name]
            all_rows.append((name, cnt, per, pts))
        else:
            all_rows.append((name, 0, 0, 0))

    total_deduct = sum(r[3] for r in all_rows)
    total_score = max(0, TOTAL - total_deduct)
    return total_score, TOTAL, all_rows


# (以下 `write_shift_result` からの処理は既存のままなので省略せず残しています)
def write_shift_result(result, staff, shuunin_list, unit_map, cont_map, role_map,
                       days_norm, requests, ab_unit_result, shuunin_unit_result,
                       kanmu_map, input_path, output_path, prev_month=None,
                       nmin_map=None, nmax_map=None, consec_night_map=None,
                       holiday_periods=None, role_map_extra=None,
                       ojt_list=None, ojt_instructor=None):
    from openpyxl import Workbook
    from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.worksheet.page import PageMargins

    if prev_month is None:
        prev_month = {}

    period_end_col_offset = None  
    try:
        from openpyxl import load_workbook as _lw
        _wb = _lw(input_path, keep_vba=True, data_only=True)
        _ws = _wb['Settings']
        _b5 = _ws['B5'].value
        if _b5 is not None:
            import pandas as _pd
            _end_date = _pd.to_datetime(_b5).to_pydatetime().replace(
                tzinfo=None, hour=0, minute=0, second=0, microsecond=0)
            for _i, _d in enumerate(days_norm):
                if _d.date() == _end_date.date():
                    period_end_col_offset = _i
                    break
    except Exception:
        pass

    result_mod = {}
    all_disp_tmp = shuunin_list + staff
    for s in all_disp_tmp:
        result_mod[s] = dict(result.get(s, {}))

    wb = Workbook()
    ws = wb.active
    ws.title = "shift_result"

    N = len(days_norm)
    weekday_ja = ["月", "火", "水", "木", "金", "土", "日"]
    DATE_START_COL = 2
    SUMMARY_COL    = DATE_START_COL + N

    SUMM_ABBR  = ["ハ", "ニ", "オ", "夜勤", "", "", "", "", "", "年", ""]
    SUMM_FULL  = ["早出", "日勤", "遅出", "夜勤", "計", "日", "土", "公", "計", "年休", "合計"]
    NUM_SUMM   = len(SUMM_FULL)

    thin   = Side(style="thin")
    medium = Side(style="medium")
    thin_border   = Border(left=thin,   right=thin,   top=thin,   bottom=thin)
    header_border = Border(left=thin,   right=thin,   top=medium, bottom=medium)

    _ojt = ojt_list or []
    _ojt_instr = ojt_instructor or {}

    # OJTをユニット最下行に配置：非OJTを先に、OJTを後に
    all_staff_ordered = shuunin_list + staff
    group1_non_ojt = [s for s in all_staff_ordered if unit_map.get(s, "") != "B" and s not in _ojt]
    group1_ojt     = [s for s in all_staff_ordered if unit_map.get(s, "") != "B" and s in _ojt]
    group1 = group1_non_ojt + group1_ojt
    group2_non_ojt = [s for s in all_staff_ordered if unit_map.get(s, "") == "B" and s not in _ojt]
    group2_ojt     = [s for s in all_staff_ordered if unit_map.get(s, "") == "B" and s in _ojt]
    group2 = group2_non_ojt + group2_ojt

    STAFF_START_ROW = 5
    first_group2_row = STAFF_START_ROW + len(group1)
    LAST_STAFF_ROW   = STAFF_START_ROW + len(group1) + len(group2) - 1
    DAILY_ROW_BASE = LAST_STAFF_ROW + 2

    SHIFT_ABBR = {"早": "ハ", "遅": "オ", "日": "ニ", "有": "年"}
    # 内部記号 → 表示記号マッピング
    DISPLAY_MAP = {"×": "土", "○": "日", "公": "公", "△": "日"}

    def display_val(s, d):
        sh = result_mod[s].get(d, "×")
        if sh in DISPLAY_MAP:
            return DISPLAY_MAP[sh]
        if sh == "日":
            return "ニ"
        if sh == "有":
            return "年"
        if sh not in ("早", "遅"):
            return sh
        abbr = SHIFT_ABBR[sh]
        if s in shuunin_list:
            unit = shuunin_unit_result.get(s, {}).get(d)
            return (unit + abbr) if unit else abbr
        elif kanmu_map.get(s, "×") == "○":
            unit = ab_unit_result.get(s, {}).get(d)
            return (unit + abbr) if unit else abbr
        else:
            unit = unit_map.get(s, "")
            return (unit + abbr) if unit in ("A", "B") else abbr

    def cell_fill(s, d):
        date_obj = days_norm[d]
        if s in requests and date_obj in requests[s]:
            _, rtype = requests[s][date_obj]
            if rtype == "希望":
                return PINK_FILL
            elif rtype == "指定":
                return GREEN_FILL
        if s in shuunin_list:
            unit = shuunin_unit_result.get(s, {}).get(d)
            sh   = result.get(s, {}).get(d, "×")
            if sh == "早" and unit:
                return BLUE_FILL
        return None

    month_label = f"{days_norm[0].month}月" if days_norm else ""
    label_col = DATE_START_COL + N // 2
    c = ws.cell(1, label_col, month_label)
    c.font = Font(bold=True, size=14)
    c.alignment = Alignment(horizontal="center")

    ws.cell(2, 1, "日").alignment = Alignment(horizontal="center")
    ws.cell(2, 1).font = Font(bold=True)
    for i, d in enumerate(days_norm):
        col = DATE_START_COL + i
        c = ws.cell(2, col, d.day)
        c.alignment = Alignment(horizontal="center")
        c.font = Font(bold=True)
        c.border = thin_border

    ws.cell(3, 1, "曜日").alignment = Alignment(horizontal="center")
    ws.cell(3, 1).font = Font(bold=True)
    for i, d in enumerate(days_norm):
        col = DATE_START_COL + i
        wd  = weekday_ja[d.weekday()]
        c   = ws.cell(3, col, wd)
        c.alignment = Alignment(horizontal="center")
        c.border = thin_border
        if d.weekday() == 5:
            c.fill = PatternFill("solid", fgColor="CCE5FF")
        elif d.weekday() == 6:
            c.fill = PatternFill("solid", fgColor="FFCCCC")

    for k, abbr in enumerate(SUMM_ABBR):
        if abbr:
            c = ws.cell(3, SUMMARY_COL + k, abbr)
            c.alignment = Alignment(horizontal="center")
            c.font = Font(bold=True)
            c.fill = YELLOW_FILL
            c.border = thin_border

    c4 = ws.cell(4, 1, "会議・委員会")
    c4.alignment = Alignment(horizontal="center", vertical="center")
    c4.font = Font(bold=True)
    c4.border = thin_border
    for i in range(N):
        ws.cell(4, DATE_START_COL + i).border = thin_border

    for k, h in enumerate(SUMM_FULL):
        c = ws.cell(4, SUMMARY_COL + k, h)
        c.alignment = Alignment(horizontal="center")
        c.font = Font(bold=True)
        c.fill = YELLOW_FILL
        c.border = header_border

    OJT_FILL = PatternFill("solid", fgColor="FFF0E0")  # OJT薄オレンジ

    def write_staff_row(row, s, extra_top=False, extra_bottom=False):
        u = unit_map.get(s, "")
        is_shuunin = (s in shuunin_list)
        is_ojt = (s in _ojt)

        # OJTは名前に指導者を併記
        display_name = s
        if is_ojt and s in _ojt_instr:
            display_name = f"{s}(OJT/{_ojt_instr[s]})"

        nc = ws.cell(row, 1, display_name)
        nc.alignment = Alignment(horizontal="center", vertical="center")
        if is_shuunin:
            nc.fill = BLUE_FILL
            nc.font = Font(bold=True)
        elif is_ojt:
            nc.fill = OJT_FILL
            nc.font = Font(italic=True)
        elif u == "A":
            nc.fill = A_UNIT_FILL
        elif u == "B":
            nc.fill = B_UNIT_FILL

        for d in range(N):
            col  = DATE_START_COL + d
            val  = display_val(s, d)
            cell = ws.cell(row, col, val)
            cell.alignment = Alignment(horizontal="center")
            f = cell_fill(s, d)
            if f:
                cell.fill = f
            # 土=青字、日=赤字、公=紫字
            if val == "土":
                cell.font = Font(color="0070C0", bold=False)
            elif val == "日":
                cell.font = Font(color="FF0000", bold=False)
            elif val == "公":
                cell.font = Font(color="7030A0", bold=False)

        ds  = get_column_letter(DATE_START_COL)
        de  = get_column_letter(DATE_START_COL + N - 1)
        rng = f"{ds}{row}:{de}{row}"
        ws.cell(row, SUMMARY_COL,     f'=COUNTIF({rng},"Aハ")+COUNTIF({rng},"Bハ")')
        ws.cell(row, SUMMARY_COL + 1, f'=COUNTIF({rng},"ニ")')
        ws.cell(row, SUMMARY_COL + 2, f'=COUNTIF({rng},"Aオ")+COUNTIF({rng},"Bオ")')
        ws.cell(row, SUMMARY_COL + 3, f'=COUNTIF({rng},"夜")')
        k_col  = SUMMARY_COL + 4
        hk     = ws.cell(row, k_col)
        hk.value = (f'={get_column_letter(SUMMARY_COL)}{row}'
                    f'+{get_column_letter(SUMMARY_COL+1)}{row}'
                    f'+{get_column_letter(SUMMARY_COL+2)}{row}'
                    f'+{get_column_letter(SUMMARY_COL+3)}{row}')
        ws.cell(row, SUMMARY_COL + 5, f'=COUNTIF({rng},"日")')
        ws.cell(row, SUMMARY_COL + 6, f'=COUNTIF({rng},"土")')
        ws.cell(row, SUMMARY_COL + 7, f'=COUNTIF({rng},"公")')
        ws.cell(row, SUMMARY_COL + 8,
                f'={get_column_letter(SUMMARY_COL+5)}{row}'
                f'+{get_column_letter(SUMMARY_COL+6)}{row}'
                f'+{get_column_letter(SUMMARY_COL+7)}{row}')
        ws.cell(row, SUMMARY_COL + 9, f'=COUNTIF({rng},"年")')
        ws.cell(row, SUMMARY_COL + 10,
                f'={get_column_letter(SUMMARY_COL+4)}{row}'
                f'+{get_column_letter(SUMMARY_COL+8)}{row}'
                f'+{get_column_letter(SUMMARY_COL+9)}{row}')
        for k2 in range(NUM_SUMM):
            c = ws.cell(row, SUMMARY_COL + k2)
            c.alignment = Alignment(horizontal="center")
            c.fill = YELLOW_FILL

        top_side    = medium if extra_top    else thin
        bottom_side = medium if extra_bottom else thin

        ws.cell(row, 1).border = Border(
            left=medium, right=thin, top=top_side, bottom=bottom_side)
        for d in range(N):
            col = DATE_START_COL + d
            ws.cell(row, col).border = Border(
                left=thin, right=thin, top=top_side, bottom=bottom_side)
        for k2 in range(NUM_SUMM):
            ws.cell(row, SUMMARY_COL + k2).border = Border(
                left=thin, right=thin, top=top_side, bottom=bottom_side)

    for idx, s in enumerate(group1):
        row = STAFF_START_ROW + idx
        is_last  = (idx == len(group1) - 1) and len(group2) > 0
        write_staff_row(row, s, extra_bottom=is_last)

    for idx, s in enumerate(group2):
        row = first_group2_row + idx
        is_first = (idx == 0)
        write_staff_row(row, s, extra_top=is_first)

    if period_end_col_offset is not None:
        period_end_col = DATE_START_COL + period_end_col_offset
        next_col       = period_end_col + 1
        apply_rows = list(range(2, LAST_STAFF_ROW + 1)) + list(range(DAILY_ROW_BASE, DAILY_ROW_BASE + 6))
        for row in apply_rows:
            c = ws.cell(row, period_end_col)
            old_b = c.border
            c.border = Border(
                left  = old_b.left  if old_b.left  else thin,
                right = medium,
                top   = old_b.top   if old_b.top   else thin,
                bottom= old_b.bottom if old_b.bottom else thin)
            if next_col <= SUMMARY_COL - 1:
                c2 = ws.cell(row, next_col)
                old_b2 = c2.border
                c2.border = Border(
                    left  = medium,
                    right = old_b2.right  if old_b2.right  else thin,
                    top   = old_b2.top    if old_b2.top    else thin,
                    bottom= old_b2.bottom if old_b2.bottom else thin)

    # 日休み・土休み・公休 用フィル
    NICHI_FILL = PatternFill("solid", fgColor="FFCCCC")  # 日=薄赤
    DOYOU_FILL = PatternFill("solid", fgColor="CCE5FF")  # 土=薄青
    KOU_FILL   = PatternFill("solid", fgColor="E8D5F5")  # 公=薄紫

    daily_labels = ["A早出", "B早出", "A遅出", "B遅出", "夜勤", "日勤"]
    daily_codes  = ["Aハ",   "Bハ",   "Aオ",  "Bオ",  "夜",  "ニ"]
    daily_fills  = [A_UNIT_FILL, B_UNIT_FILL, A_UNIT_FILL, B_UNIT_FILL, GRAY_FILL, GRAY_FILL]
    daily_font_colors = [None, None, None, None, None, None]

    cnt_start_row = STAFF_START_ROW
    cnt_end_row   = LAST_STAFF_ROW

    for k, (lbl, dv, fill) in enumerate(zip(daily_labels, daily_codes, daily_fills)):
        dr = DAILY_ROW_BASE + k
        c = ws.cell(dr, 1, lbl)
        c.fill = fill
        c.alignment = Alignment(horizontal="center")
        c.font = Font(bold=True)
        c.border = Border(left=medium, right=thin, top=thin, bottom=thin)

        for i in range(N):
            col = DATE_START_COL + i
            col_letter = get_column_letter(col)
            cnt_range = f"{col_letter}{cnt_start_row}:{col_letter}{cnt_end_row}"
            formula_val = f'=COUNTIF({cnt_range},"{dv}")'
            dc = ws.cell(dr, col, formula_val)
            dc.alignment = Alignment(horizontal="center")
            dc.border = thin_border

        for k2 in range(NUM_SUMM):
            ws.cell(dr, SUMMARY_COL + k2).border = thin_border

    red_font_rule = CellIsRule(
        operator="equal",
        formula=["0"],
        font=Font(color="FF0000", bold=True))
    first_date_letter = get_column_letter(DATE_START_COL)
    last_date_letter  = get_column_letter(DATE_START_COL + N - 1)
    for k in range(len(daily_labels)):
        dr = DAILY_ROW_BASE + k
        range_str = f"{first_date_letter}{dr}:{last_date_letter}{dr}"
        ws.conditional_formatting.add(range_str, red_font_rule)

    sep_row = LAST_STAFF_ROW + 1
    for col in range(1, SUMMARY_COL + NUM_SUMM):
        ws.cell(sep_row, col).border = thin_border

    ws.cell(1, 1).border = Border(left=medium, right=thin, top=thin, bottom=thin)
    for i in range(N):
        ws.cell(1, DATE_START_COL + i).border = thin_border

    for r in [2, 3, 4]:
        ws.cell(r, 1).border = Border(left=medium, right=thin, top=thin, bottom=thin)

    ws.column_dimensions["A"].width = 10
    for i in range(N):
        ws.column_dimensions[get_column_letter(DATE_START_COL + i)].width = 5
    for k in range(NUM_SUMM):
        ws.column_dimensions[get_column_letter(SUMMARY_COL + k)].width = 7

    for r in range(1, 5):
        ws.row_dimensions[r].height = 18
    for r in range(STAFF_START_ROW, LAST_STAFF_ROW + 1):
        ws.row_dimensions[r].height = 16
    for k in range(len(daily_labels)):
        ws.row_dimensions[DAILY_ROW_BASE + k].height = 16

    from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize   = 9
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(
        left=0.4, right=0.4, top=0.7, bottom=0.7, header=0.3, footer=0.3)
    ws.print_title_rows = "1:4"

    # ── スコアシート ────────────────────────────────────────────────
    if nmin_map and nmax_map and consec_night_map is not None and holiday_periods is not None:
        total_score, total_max, deductions = score_shift(
            result, staff, shuunin_list, days_norm, requests,
            prev_month, cont_map, role_map, nmin_map, nmax_map,
            consec_night_map, holiday_periods, unit_map,
            ab_unit_result, shuunin_unit_result, kanmu_map
        )

        ws_score = wb.create_sheet("スコアレポート")
        SCORE_FILLS = {
            "header": PatternFill("solid", fgColor="1F4E79"),
            "ok":     PatternFill("solid", fgColor="E2EFDA"),
            "warn":   PatternFill("solid", fgColor="FFF2CC"),
            "bad":    PatternFill("solid", fgColor="FCE4D6"),
            "total":  PatternFill("solid", fgColor="D6E4F0"),
        }

        # タイトル行
        c = ws_score.cell(1, 1, "シフト表 自己採点レポート")
        c.font = Font(bold=True, size=14, color="FFFFFF")
        c.fill = SCORE_FILLS["header"]
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws_score.merge_cells("A1:E1")
        ws_score.row_dimensions[1].height = 28

        # 総合スコア
        score_pct = total_score / total_max * 100
        score_color = "ok" if score_pct >= 80 else "warn" if score_pct >= 60 else "bad"
        ws_score.cell(2, 1, "総合スコア").font = Font(bold=True)
        ws_score.cell(2, 1).alignment = Alignment(horizontal="center")
        ws_score.cell(2, 1).fill = SCORE_FILLS[score_color]
        c_score = ws_score.cell(2, 2, f"{total_score} / {total_max} 点")
        c_score.font = Font(bold=True, size=14,
                            color="375623" if score_pct >= 80 else "7F4B00" if score_pct >= 60 else "7F1810")
        c_score.alignment = Alignment(horizontal="left", vertical="center")
        ws_score.row_dimensions[2].height = 24

        # 減点なしの場合も全項目を出力
        # ヘッダー行
        headers = ["採点項目", "件数", "1件あたり減点", "合計減点", "評価"]
        for col, h in enumerate(headers, 1):
            c = ws_score.cell(3, col, h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = SCORE_FILLS["header"]
            c.alignment = Alignment(horizontal="center")
            c.border = thin_border

        # 全採点項目の明細（減点0も含む）
        for row_i, (name, cnt, per, total_d) in enumerate(deductions, 4):
            if total_d == 0:
                fill_key = "ok"
                font_color = "375623"
            elif total_d <= 30:
                fill_key = "warn"
                font_color = "7F4B00"
            else:
                fill_key = "bad"
                font_color = "C00000"

            ws_score.cell(row_i, 1, name).alignment = Alignment(horizontal="left")
            ws_score.cell(row_i, 1).fill = SCORE_FILLS[fill_key]
            ws_score.cell(row_i, 1).border = thin_border

            ws_score.cell(row_i, 2, cnt).alignment = Alignment(horizontal="center")
            ws_score.cell(row_i, 2).border = thin_border

            ws_score.cell(row_i, 3, f"-{per}点").alignment = Alignment(horizontal="center")
            ws_score.cell(row_i, 3).border = thin_border

            if total_d == 0:
                ws_score.cell(row_i, 4, "±0").alignment = Alignment(horizontal="center")
                ws_score.cell(row_i, 4).font = Font(color="375623")
            else:
                ws_score.cell(row_i, 4, f"-{total_d}点").alignment = Alignment(horizontal="center")
                ws_score.cell(row_i, 4).font = Font(bold=True, color=font_color)
            ws_score.cell(row_i, 4).border = thin_border

            star = "★★★" if total_d >= 100 else "★★" if total_d >= 50 else "★" if total_d >= 20 else ("✓" if total_d == 0 else "")
            ws_score.cell(row_i, 5, star).alignment = Alignment(horizontal="center")
            ws_score.cell(row_i, 5).font = Font(color="375623" if total_d == 0 else "C00000")
            ws_score.cell(row_i, 5).border = thin_border

        # 合計行
        last_row = 4 + len(deductions)
        ws_score.cell(last_row, 1, "合計減点").font = Font(bold=True)
        ws_score.cell(last_row, 1).fill = SCORE_FILLS["total"]
        ws_score.cell(last_row, 4, f"-{total_max - total_score}点").font = Font(bold=True, color="C00000")
        ws_score.cell(last_row, 4).fill = SCORE_FILLS["total"]
        for col in range(1, 6):
            ws_score.cell(last_row, col).border = thin_border

        # 列幅
        ws_score.column_dimensions["A"].width = 32
        ws_score.column_dimensions["B"].width = 10
        ws_score.column_dimensions["C"].width = 16
        ws_score.column_dimensions["D"].width = 12
        ws_score.column_dimensions["E"].width = 8

        # スコアを右下の空欄エリア（日別集計行 × サマリー列）に表示
        sc = SUMMARY_COL
        sr = DAILY_ROW_BASE      # 開始行
        er = DAILY_ROW_BASE + len(daily_labels) - 1  # 終了行

        # 結合セル: 全サマリー列 × 全日別集計行
        score_range = f"{get_column_letter(sc)}{sr}:{get_column_letter(sc + NUM_SUMM - 1)}{er}"
        ws.merge_cells(score_range)
        score_cell = ws.cell(sr, sc)
        score_cell.value = f"採点\n{total_score} / {total_max} 点"
        score_cell.font = Font(
            bold=True, size=16,
            color="375623" if score_pct >= 80 else "7F4B00" if score_pct >= 60 else "C00000")
        score_cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)
        score_cell.fill = PatternFill(
            "solid",
            fgColor="E2EFDA" if score_pct >= 80 else "FFF2CC" if score_pct >= 60 else "FCE4D6")
        score_cell.border = Border(
            left=thin, right=thin, top=thin, bottom=thin)

    wb.save(output_path)


# ========================================================
# Web UI HTML
# ========================================================
with open(os.path.join(os.path.dirname(os.path.abspath(__file__)), "index.html"), encoding="utf-8") as _f:
    HTML_CONTENT = _f.read()



# ========================================================
# FastAPI Routes
# ========================================================
@app.get("/", response_class=HTMLResponse)
async def index():
    return HTMLResponse(content=HTML_CONTENT)

@app.get("/health")
async def health():
    return {"status": "ok", "version": "6.0"}

@app.get("/start-session")
async def start_session():
    """計算開始前にUIDを発行してSSE購読の準備をする"""
    uid = str(uuid.uuid4())
    _get_progress_queue(uid)
    return {"uid": uid}


@app.get("/progress/{uid}")
async def progress_stream(uid: str):
    """SSE: 計算進捗をリアルタイムストリーミング"""
    async def event_gen():
        q = _progress_queues.get(uid)
        if q is None:
            yield "data: (進捗なし)\n\n"
            return
        while True:
            try:
                msg = await asyncio.get_event_loop().run_in_executor(
                    None, lambda: q.get(timeout=120))
                if msg is None:
                    yield "data: [完了]\n\n"
                    break
                yield f"data: {msg}\n\n"
            except Exception:
                break
    return StreamingResponse(event_gen(), media_type="text/event-stream",
                             headers={"Cache-Control": "no-cache",
                                      "X-Accel-Buffering": "no"})


@app.get("/preview-result/{uid}")
async def preview_result(uid: str):
    """生成済みシフト結果のプレビューJSONを返す（再計算なし）
    単一パターン: uid そのまま
    複数パターン: uid_p1, uid_p2... のいずれか、または uid_count でパターン数を返す
    """
    # パターン数問い合わせ
    if uid.endswith("_count"):
        base = uid[:-6]
        count = _result_cache.get(uid)
        if count is None:
            raise HTTPException(status_code=404, detail="データがありません")
        return {"count": count}
    data = _result_cache.get(uid)
    if data is None:
        raise HTTPException(status_code=404, detail="プレビューデータがありません。先に生成を実行してください。")
    return data


def _suggest_relaxation(error_msg: str) -> list:
    """INFEASIBLE エラーから緩和できる制約を提案する"""
    suggestions = []
    e = error_msg

    if "Aユニット早出" in e or "Bユニット早出" in e:
        suggestions.append("→ 主任をAまたはBユニット兼務に設定すると早出補充に使えます（Staff_MasterのユニットをA・Bに変更）")
    if "遅出" in e and "配置できるスタッフ" in e:
        suggestions.append("→ 該当日の希望休を取り消すか、備考の勤務制限（遅出禁止等）を緩めてください")
    if "夜勤に配置できる" in e:
        suggestions.append("→ 夜勤可能スタッフの最高回数を1〜2回増やすか、夜勤不可スタッフの希望休を見直してください")
    if "夜勤最少数合計" in e:
        suggestions.append("→ 夜勤最少数の合計を対象日数以下に下げてください（例：各スタッフの最少数を1減らす）")
    if "夜勤最高数合計" in e:
        suggestions.append("→ 夜勤最高数の合計を対象日数以上に増やしてください（例：各スタッフの最高数を1増やす）")
    if "公休数" in e or "希望休の数" in e:
        suggestions.append("→ Shift_Requestsの希望休を減らすか、Settingsの公休日数設定を増やしてください")
    if "夜勤上限" in e and "公休数" in e:
        suggestions.append("→ 夜勤上限回数を公休数より少なくするか、公休数を夜勤上限以上に増やしてください")
    if "連続夜勤" in e:
        suggestions.append("→ 連続夜勤希望◯のスタッフを増やすか、夜勤最少数を下げてください")
    if "前半夜勤NG" in e or "15日以前" in e:
        suggestions.append("→ 前半夜勤NGのスタッフの夜勤最少数を下げるか、NGを外してください")
    if "主任を使った補充を含めても" in e:
        suggestions.append("→ 以下のいずれかを試してください：")
        suggestions.append("   ① 最も制約の多い日の希望休・指定勤務を1つ取り消す")
        suggestions.append("   ② 夜勤最少数を各スタッフ1回ずつ下げる")
        suggestions.append("   ③ 備考の勤務制限（早出のみ・遅出のみ等）を一時的に外す")
        suggestions.append("   ④ 公休日数の設定を1〜2日増やす（Settingsシート）")

    if "OJT" in e or "(OJT)" in e:
        suggestions.append("→ OJTスタッフの希望休・指定勤務が原因の可能性があります：")
        suggestions.append("   ① OJTの希望休を減らすか、公休数設定を見直してください")
        suggestions.append("   ② OJTと指導者が同じ日に休みを入れていないか確認してください")
        suggestions.append("   ③ OJTの固定公休と指定勤務が衝突していないか確認してください")

    if not suggestions:
        suggestions.append("→ 希望シフト・夜勤回数・公休数の設定を見直してください")
        suggestions.append("→ 事前チェックボタンで詳細な矛盾箇所を確認できます")

    return suggestions


@app.post("/validate")
async def validate(file: UploadFile = File(...)):
    """高速フィジビリティチェック: 15秒以内に実現可能解が存在するか確認する"""
    uid = str(uuid.uuid4())
    orig_name = file.filename or "upload.xlsx"
    ext = os.path.splitext(orig_name)[1].lower()
    if ext not in [".xlsx", ".xls", ".xlsm"]:
        ext = ".xlsx"
    in_p = os.path.join(TEMP_DIR, f"in_{uid}{ext}")
    try:
        with open(in_p, "wb") as f:
            shutil.copyfileobj(file.file, f)
        result = generate_shift(in_p, validate_only=True)
        return result
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        try: os.remove(in_p)
        except: pass


@app.post("/generate-shift")
async def generate(file: UploadFile = File(...), patterns: int = 1, uid: str = ""):
    if not uid:
        uid = str(uuid.uuid4())
    if uid not in _progress_queues:
        _get_progress_queue(uid)
    orig_name = file.filename or "upload.xlsx"
    ext  = os.path.splitext(orig_name)[1].lower()
    if ext not in [".xlsx", ".xls", ".xlsm"]:
        ext = ".xlsx"
    in_p = os.path.join(TEMP_DIR, f"in_{uid}{ext}")
    patterns = max(1, min(patterns, 5))

    try:
        with open(in_p, "wb") as f:
            shutil.copyfileobj(file.file, f)

        _push_progress(uid, f"ファイル読み込み完了: {orig_name}")

        if patterns == 1:
            out_p = os.path.join(TEMP_DIR, f"out_{uid}.xlsx")
            _push_progress(uid, "制約モデルを構築中...")
            (result, staff, shuunin_list, unit_map, cont_map, role_map,
             days_norm, requests, ab_unit_result, shuunin_unit_result,
             kanmu_map, prev_month, nmin_map, nmax_map,
             consec_night_map, holiday_periods,
             ojt_list, ojt_instructor) = generate_shift(
                 in_p, random_seed=0, timeout=300, progress_uid=uid)
            _push_progress(uid, "Excelファイルを生成中...")
            write_shift_result(
                result, staff, shuunin_list, unit_map, cont_map, role_map,
                days_norm, requests, ab_unit_result, shuunin_unit_result,
                kanmu_map, in_p, out_p, prev_month=prev_month,
                nmin_map=nmin_map, nmax_map=nmax_map,
                consec_night_map=consec_night_map, holiday_periods=holiday_periods,
                ojt_list=ojt_list, ojt_instructor=ojt_instructor)
            # 生成結果をプレビュー用にキャッシュ
            _cache_preview(uid, result, staff, shuunin_list, unit_map, cont_map,
                           role_map, days_norm, requests, ab_unit_result, shuunin_unit_result,
                           kanmu_map, prev_month, nmin_map, nmax_map,
                           consec_night_map, holiday_periods)
            _close_progress(uid)
            return FileResponse(
                out_p, filename="SS_Result.xlsx",
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"X-Progress-UID": uid})
        else:
            # 複数パターン → 並列実行してZIPで返す
            # 並列実行: 全パターンを同時に走らせ、合計時間を300秒に抑える
            per_timeout = 300  # 並列なので全パターン同時実行 → 合計300秒
            import zipfile, io
            from concurrent.futures import ThreadPoolExecutor, as_completed

            seeds = [0, 42, 123, 777, 999]

            # 並列実行: CPU競合を防ぐためワーカー数をパターン数で分割
            import os as _os
            total_cpu = _os.cpu_count() or 8
            per_workers = max(1, total_cpu // patterns)
            per_timeout = 300

            _push_progress(uid, f"{patterns}パターンを並列生成中（同時実行・各最大{per_timeout}秒・各{per_workers}スレッド）...")

            def run_one(i):
                """1パターン生成して (i, out_path or None, error or None) を返す"""
                seed = seeds[i] if i < len(seeds) else i * 137
                out_p = os.path.join(TEMP_DIR, f"out_{uid}_p{i+1}.xlsx")
                _push_progress(uid, f"パターン {i+1} 開始（seed={seed}）...")
                try:
                    (res, st, sh_l, u_map, c_map, r_map,
                     d_norm, reqs, ab_ur, sh_ur,
                     k_map, p_month, nm_min, nm_max,
                     cn_map, h_periods,
                     o_list, o_instr) = generate_shift(
                         in_p, random_seed=seed, timeout=per_timeout,
                         progress_uid=uid, num_workers=per_workers)
                    write_shift_result(
                        res, st, sh_l, u_map, c_map, r_map,
                        d_norm, reqs, ab_ur, sh_ur,
                        k_map, in_p, out_p, prev_month=p_month,
                        nmin_map=nm_min, nmax_map=nm_max,
                        consec_night_map=cn_map, holiday_periods=h_periods,
                        ojt_list=o_list, ojt_instructor=o_instr)
                    # 各パターンをプレビューキャッシュに保存
                    _cache_preview(f"{uid}_p{i+1}", res, st, sh_l, u_map, c_map,
                                   r_map, d_norm, reqs, ab_ur, sh_ur,
                                   k_map, p_month, nm_min, nm_max, cn_map, h_periods)
                    _push_progress(uid, f"✓ パターン {i+1} 完了")
                    return (i, out_p, None)
                except Exception as e:
                    _push_progress(uid, f"✗ パターン {i+1} 失敗: {e}")
                    return (i, None, str(e))

            zip_buf = io.BytesIO()
            generated = 0
            errors = []
            out_paths = {}

            with ThreadPoolExecutor(max_workers=patterns) as executor:
                futures = {executor.submit(run_one, i): i for i in range(patterns)}
                for future in as_completed(futures):
                    i, out_p, err = future.result()
                    if err:
                        errors.append(f"パターン{i+1}: {err}")
                    else:
                        out_paths[i] = out_p

            # 番号順にZIPへ追加
            with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                for i in sorted(out_paths.keys()):
                    out_p = out_paths[i]
                    zf.write(out_p, f"SS_Result_パターン{i+1}.xlsx")
                    generated += 1
                    try: os.remove(out_p)
                    except: pass

            if generated == 0:
                raise Exception("\n".join(errors))

            # 生成成功パターン数をキャッシュに記録（フロントがタブ数を知るため）
            _result_cache[f"{uid}_count"] = generated

            _push_progress(uid, f"{patterns}パターンを並列生成中（同時実行・各最大{per_timeout}秒）...")

            zip_buf = io.BytesIO()
            generated = 0
            errors = []
            out_paths = {}

            with ThreadPoolExecutor(max_workers=patterns) as executor:
                futures = {executor.submit(run_one, i): i for i in range(patterns)}
                for future in as_completed(futures):
                    i, out_p, err = future.result()
                    if err:
                        errors.append(f"パターン{i+1}: {err}")
                    else:
                        out_paths[i] = out_p

            # 番号順にZIPへ追加
            with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                for i in sorted(out_paths.keys()):
                    out_p = out_paths[i]
                    zf.write(out_p, f"SS_Result_パターン{i+1}.xlsx")
                    generated += 1
                    try: os.remove(out_p)
                    except: pass

            if generated == 0:
                raise Exception("\n".join(errors))

            zip_buf.seek(0)
            zip_path = os.path.join(TEMP_DIR, f"out_{uid}.zip")
            with open(zip_path, "wb") as f:
                f.write(zip_buf.getvalue())
            _push_progress(uid, f"全{generated}パターン完了！ZIPをダウンロードします。")
            _close_progress(uid)
            return FileResponse(
                zip_path, filename="SS_Result_複数パターン.zip",
                media_type="application/zip",
                headers={"X-Progress-UID": uid})

    except Exception as e:
        import traceback; traceback.print_exc()
        _close_progress(uid)
        suggestions = _suggest_relaxation(str(e))
        detail = str(e)
        if suggestions:
            detail += "\n\n【緩和提案】\n" + "\n".join(suggestions)
        raise HTTPException(status_code=500, detail=detail)
    finally:
        try: os.remove(in_p)
        except: pass


# ========================================================
# スタンドアロン起動
# ========================================================
if __name__ == "__main__":
    import uvicorn, webbrowser, threading, time

    def open_browser():
        time.sleep(2.0)
        webbrowser.open("http://localhost:8000")

    port = int(os.environ.get("PORT", 8000))
    host = os.environ.get("HOST", "0.0.0.0")
    if os.environ.get("AUTO_BROWSER", "1") == "1" and port == 8000:
        threading.Thread(target=open_browser, daemon=True).start()

    print("=" * 50)
    print(" Smart Shift by OR-Tools")
    print(f" http://localhost:{port}")
    print("=" * 50)
    uvicorn.run("main:app", host=host, port=port, reload=False)
